import streamlit as st
import pandas as pd
import io # BytesIO 사용을 위해
import openpyxl # <--- 직접 사용을 위해 추가
import tempfile # 임시 파일 생성을 위해 추가
import os       # 임시 파일 삭제를 위해 추가
import re       # 정규 표현식을 위해 추가

# create_excel.py에서 함수 임포트
# 이 파일이 streamlit_app.py와 같은 디렉토리에 있다고 가정
try:
    from create_excel import extract_specific_data
except ImportError:
    st.error("오류: 'create_excel.py' 파일을 찾을 수 없거나, 파일 내에 'extract_specific_data' 함수가 없습니다. 동일한 디렉토리에 해당 파일이 있는지 확인해주세요.")
    st.stop()


def generate_excel_for_download(all_rows_list):
    """
    모든 파일에서 처리된 행들의 리스트를 받아 단일 시트를 가진 Excel 파일을 생성합니다.
    """
    output = io.BytesIO()
    
    # 컬럼 헤더 수정
    headers = ["위치", "계측기명", "현재값"]
    final_df = pd.DataFrame()

    if all_rows_list:
        try:
            final_df = pd.DataFrame(all_rows_list, columns=headers)
            # nan 값을 가진 행 제거
            final_df = final_df[final_df["현재값"].str.lower() != "nan"]
            # 위치 기준으로 오름차순 정렬
            final_df = final_df.sort_values(by="위치", ascending=True)
        except Exception as e:
            print(f"종합 결과 DataFrame 생성 중 오류: {e}")

    if not final_df.empty:
        # 1단계: Pandas를 사용하여 DataFrame을 BytesIO에 씁니다.
        final_df.to_excel(output, sheet_name='종합 결과', index=False, header=True, engine='openpyxl')
        output.seek(0)

        # 2단계: openpyxl을 사용하여 BytesIO에서 워크북을 로드하고 셀 병합
        try:
            workbook = openpyxl.load_workbook(output)
            sheet = workbook['종합 결과']

            # 모든 열의 너비를 30으로 설정
            for col in range(1, sheet.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 30

            # '위치' 열 (A열) 셀 병합 로직
            start_row = 2  # 첫 번째 행은 헤더이므로 2번째 행부터 시작
            current_value = None
            merge_start_row = 2

            for row_idx in range(start_row, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=1).value
                
                # 현재 값이 이전 값과 다르거나 마지막 행인 경우
                if cell_value != current_value or row_idx == sheet.max_row:
                    # 이전 그룹에 대한 병합 처리 (2개 이상의 행이 있는 경우에만)
                    if current_value is not None and row_idx - merge_start_row > 0:
                        end_row = row_idx - 1 if cell_value != current_value else row_idx
                        sheet.merge_cells(start_row=merge_start_row, start_column=1, 
                                        end_row=end_row, end_column=1)
                        
                        # 병합된 셀의 수직 가운데 정렬
                        merged_cell = sheet.cell(row=merge_start_row, column=1)
                        merged_cell.alignment = openpyxl.styles.Alignment(vertical='center')
                    
                    # 새로운 그룹 시작
                    current_value = cell_value
                    merge_start_row = row_idx

            # 수정된 워크북을 새로운 BytesIO 객체에 저장
            final_processed_output = io.BytesIO()
            workbook.save(final_processed_output)
            final_processed_output.seek(0)
            processed_data = final_processed_output.getvalue()
            return processed_data
            
        except Exception as e:
            print(f"Excel 파일 셀 병합 중 오류: {e}")
            output.seek(0)
            processed_data = output.getvalue()
            return processed_data
        
    else: # final_df가 비어있는 경우
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            pd.DataFrame({'정보':['처리할 데이터가 없거나, 종합 결과 생성에 실패했습니다.']}).to_excel(writer, sheet_name='정보', index=False)
        processed_data = output.getvalue()
        return processed_data

def main():
    st.set_page_config(layout="wide", page_title="엑셀 데이터 처리기")

    # 세션 상태 초기화
    if 'processed_files_count' not in st.session_state: st.session_state.processed_files_count = 0
    if 'failed_files_info' not in st.session_state: st.session_state.failed_files_info = []
    if 'all_accumulated_rows' not in st.session_state: st.session_state.all_accumulated_rows = []
    if 'download_data' not in st.session_state: st.session_state.download_data = None
    if 'processed_files' not in st.session_state: st.session_state.processed_files = set()
    if 'last_date' not in st.session_state: st.session_state.last_date = None  # 마지막 날짜 저장용

    # --- 사이드바 ---
    with st.sidebar:
        st.title("📁 엑셀 파일 업로드")
        st.write("엑셀 파일을 선택하세요. (여러 개 선택 가능)")
        # accept_multiple_files=True 로 변경
        uploaded_file_list = st.file_uploader("Drag and drop file(s) here", type=["xlsx", "xls"], 
                                            label_visibility="collapsed", key="file_uploader", 
                                            accept_multiple_files=True)
        st.caption("Limit 200MB per file • XLSX, XLS")

        if uploaded_file_list: # 리스트 형태로 반환됨
            # 새로운 파일만 처리
            new_files = [f for f in uploaded_file_list if f.name not in st.session_state.processed_files]
            if new_files:  # 새로운 파일이 있는 경우에만 처리
                st.session_state.failed_files_info = []  # 실패 정보는 초기화 (새로운 시도의 실패만 표시)
                st.session_state.download_data = None    # 다운로드 데이터는 전체 결과로 다시 생성
                
                for i, uploaded_file_item in enumerate(new_files):
                    st.markdown(f"""--- 
     **새 파일 처리 중: {uploaded_file_item.name}**""")
                    # 개별 파일 처리 로직 (기존 로직을 여기에 적용하고 결과 누적)
                    # 아래는 단일 파일 처리 로직을 가져와 수정 적용한 부분입니다.
                    current_file_error_message = None
                    temp_file_path_item = None
                    initial_openpyxl_style_error_occurred_item = False
                    active_sheet_name_item = "Sheet1" # 기본값
                    df_item = None

                    try:
                        progress_bar_item = st.progress(0, text=f"{uploaded_file_item.name}: 임시 저장 중...")
                        original_file_name_item = uploaded_file_item.name
                        file_suffix_item = ".xlsx"
                        if '.xls' in original_file_name_item.lower():
                            if '.xlsx' in original_file_name_item.lower(): file_suffix_item = ".xlsx"
                            elif '.xlsb' in original_file_name_item.lower(): file_suffix_item = ".xlsb"
                            elif '.xlsm' in original_file_name_item.lower(): file_suffix_item = ".xlsm"
                            else: file_suffix_item = ".xls"
                        
                        with tempfile.NamedTemporaryFile(delete=False, suffix=file_suffix_item) as tmp_item:
                            tmp_item.write(uploaded_file_item.getvalue())
                            temp_file_path_item = tmp_item.name
                        progress_bar_item.progress(10, text=f"{uploaded_file_item.name}: 읽기 시도 (기본 엔진)...")

                        # --- 기존 파일 읽기 시도 로직 (pd.read_excel, calamine, CSV 변환) --- 
                        # 이 부분은 길어서 요약. temp_file_path_item, file_suffix_item, progress_bar_item 사용
                        # active_sheet_name_item 과 df_item 이 이 과정에서 결정됨.
                        # (이전 코드의 파일 읽기 로직을 여기에 통합하고, 변수명 _item 접미사 사용)
                        # [시도1: Pandas 기본 엔진]
                        try:
                            excel_file_dict = pd.read_excel(temp_file_path_item, header=0, sheet_name=None)
                            if excel_file_dict:
                                active_sheet_name_item = list(excel_file_dict.keys())[0]
                                df_item = excel_file_dict[active_sheet_name_item]
                                progress_bar_item.progress(30, text=f"{uploaded_file_item.name}: '{active_sheet_name_item}' 읽기 완료 (기본)")
                            else:
                                df_item = pd.DataFrame()
                                current_file_error_message = "Excel에서 시트를 찾을 수 없습니다."
                        except Exception as e_pandas_item:
                            if "_NamedCellStyle" in str(e_pandas_item) and "NoneType" in str(e_pandas_item):
                                initial_openpyxl_style_error_occurred_item = True
                                current_file_error_message = f"기본 엔진 실패. 'calamine' 시도: {e_pandas_item}"
                                progress_bar_item.progress(35, text=f"{uploaded_file_item.name}: 'calamine' 시도 중...")
                                if file_suffix_item in [".xlsx", ".xlsb", ".xlsm"]:
                                    try:
                                        excel_file_dict_calamine = pd.read_excel(temp_file_path_item, header=0, engine='calamine', sheet_name=None)
                                        if excel_file_dict_calamine:
                                            active_sheet_name_item = list(excel_file_dict_calamine.keys())[0]
                                            df_item = excel_file_dict_calamine[active_sheet_name_item]
                                            progress_bar_item.progress(50, text=f"{uploaded_file_item.name}: '{active_sheet_name_item}' 읽기 완료 (calamine)")
                                            current_file_error_message = None # calamine 성공 시 이전 오류 메시지 초기화
                                        else: df_item = pd.DataFrame(); current_file_error_message = "Calamine: 시트 없음"
                                    except ImportError: df_item = pd.DataFrame(); current_file_error_message = "Calamine 엔진 설치 필요"
                                    except Exception as e_cal_item: df_item = pd.DataFrame(); current_file_error_message = f"Calamine 오류: {e_cal_item}"
                                else: df_item = pd.DataFrame(); current_file_error_message = f"{file_suffix_item} Calamine 미지원"
                            else: df_item = pd.DataFrame(); current_file_error_message = f"기본 엔진 오류: {e_pandas_item}"
                        
                        # [시도3: CSV 내부 변환]
                        if (df_item is None or df_item.empty) and initial_openpyxl_style_error_occurred_item:
                            progress_bar_item.progress(55, text=f"{uploaded_file_item.name}: CSV 변환 시도...")
                            workbook_data_only_item = None
                            try:
                                workbook_data_only_item = openpyxl.load_workbook(temp_file_path_item, read_only=True, data_only=True)
                                if workbook_data_only_item.sheetnames:
                                    active_sheet_name_item = workbook_data_only_item.sheetnames[0]
                                    sheet_data_only_item = workbook_data_only_item[active_sheet_name_item]
                                    data_list_item = list(sheet_data_only_item.values)
                                    if data_list_item and data_list_item[0] is not None:
                                        df_from_data_only_item = pd.DataFrame(data_list_item[1:], columns=data_list_item[0])
                                        if not df_from_data_only_item.empty:
                                            csv_buffer_item = io.StringIO()
                                            df_from_data_only_item.to_csv(csv_buffer_item, index=False)
                                            csv_buffer_item.seek(0)
                                            df_item = pd.read_csv(csv_buffer_item)
                                            progress_bar_item.progress(70, text=f"{uploaded_file_item.name}: CSV 변환 읽기 성공")
                                            current_file_error_message = None # CSV 성공 시 이전 오류 메시지 초기화
                                        else: current_file_error_message = "CSV: 내용 비어있음"
                                    else: current_file_error_message = "CSV: 데이터 추출 불가"
                                else: current_file_error_message = "CSV: 시트 없음"
                            except Exception as e_csv_item:
                                current_file_error_message = f"CSV 변환 오류: {e_csv_item}"
                            finally:
                                if workbook_data_only_item: workbook_data_only_item.close()
                        
                        if df_item is None: df_item = pd.DataFrame() # 최종 안전장치
                        # --- 파일 읽기 시도 로직 끝 --- 

                        if not df_item.empty:
                            progress_bar_item.progress(75, text=f"{uploaded_file_item.name}: 데이터 추출 중...")
                            header_list_item = df_item.columns.tolist()
                            data_values_list_item = df_item.values.tolist()
                            data_for_extraction_item = [header_list_item] + data_values_list_item
                            
                            # 원본 데이터에서 1행 값을 직접 가져오기 (header=0으로 읽었으므로 df의 컬럼명이 1행 값임)
                            first_row_values = df_item.columns.tolist()
                            
                            extracted_results_item = extract_specific_data(data_for_extraction_item, sheet_title=active_sheet_name_item)
                            
                            # first_row_data_item은 더 이상 사용하지 않음
                            last_row_data_item = extracted_results_item.get('last_row_values')
                            
                            # 위치: 시트 제목에서 "ALL" 제거 및 공백 정리 (대소문자 구분 없이)
                            location_val = active_sheet_name_item
                            location_val = location_val.replace("ALL", "").replace("all", "").replace("All", "").strip()
                            
                            # "INC_" 뒤의 모든 문자 삭제
                            if "INC_" in location_val:
                                location_val = location_val.split("INC_")[0].strip()
                            elif "inc_" in location_val:
                                location_val = location_val.split("inc_")[0].strip()
                            elif "Inc_" in location_val:
                                location_val = location_val.split("Inc_")[0].strip()
                            
                            # 연속된 공백을 하나로 치환
                            location_val = " ".join(location_val.split())
                            # "출입구" 앞에 "도림" 추가
                            if "출입구" in location_val and not location_val.startswith("도림"):
                                location_val = location_val.replace("출입구", "도림출입구")

                            # 현재값: last_row_data_item의 각 요소에 대해 행 생성 (A열 값은 제외)
                            if isinstance(last_row_data_item, list) and last_row_data_item and len(last_row_data_item) > 1:
                                # A열의 마지막 행 값에서 날짜 추출
                                last_a_column_value = str(last_row_data_item[0])
                                # YYYY-MM-DD 형식의 날짜 찾기
                                date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', last_a_column_value)
                                if date_match:
                                    year = date_match.group(1)
                                    month = date_match.group(2).lstrip('0')  # 앞의 0 제거
                                    day = date_match.group(3).lstrip('0')    # 앞의 0 제거
                                    current_date = f"{year}년 {month}월 {day}일"
                                    
                                    # 날짜 일치 여부 확인
                                    if st.session_state.last_date is None:  # 첫 파일
                                        st.session_state.last_date = current_date
                                    elif st.session_state.last_date != current_date:  # 날짜 불일치
                                        raise ValueError(f"날짜 불일치: {st.session_state.last_date} != {current_date}")
                                else:
                                    raise ValueError(f"A열 마지막 행에서 날짜 형식(YYYY-MM-DD)을 찾을 수 없습니다.")

                                # 첫 번째 값(A열)을 제외한 나머지 값들에 대해 처리
                                for col_idx, current_value_item in enumerate(last_row_data_item[1:], 1):
                                    if col_idx < len(first_row_values):  # first_row_values의 범위 내에서만 처리
                                        instrument_name_val = str(first_row_values[col_idx])  # 해당 열의 1행 값
                                        current_value_str = str(current_value_item)
                                        # "nan" 값을 가진 행은 건너뛰기
                                        if current_value_str.lower() != "nan":
                                            new_row_for_output = [location_val, instrument_name_val, current_value_str]
                                            st.session_state.all_accumulated_rows.append(new_row_for_output)
                                
                                st.session_state.processed_files_count += 1
                                st.session_state.processed_files.add(uploaded_file_item.name)  # 처리 완료된 파일 기록
                                st.success(f"파일 '{uploaded_file_item.name}' ({location_val}) 처리 성공!")
                            else:
                                if not current_file_error_message:
                                    current_file_error_message = f"'{uploaded_file_item.name}' ({location_val}) 처리 중 현재값 데이터 추출 실패 (마지막 행 정보가 비어있거나 유효하지 않습니다)."
                        else: # df_item이 비어있는 경우
                            if not current_file_error_message:
                                current_file_error_message = "파일을 읽었으나 내용이 비어있습니다."
                        
                        progress_bar_item.empty()

                    except Exception as e_file_item:
                        current_file_error_message = f"파일 처리 중 예기치 않은 오류: {str(e_file_item)}"
                    finally:
                        if temp_file_path_item and os.path.exists(temp_file_path_item):
                            try: os.remove(temp_file_path_item)
                            except Exception as e_remove_item: st.warning(f"임시 파일 삭제 오류 ({uploaded_file_item.name}): {e_remove_item}")
                    
                    if current_file_error_message:
                        st.error(f"파일 '{uploaded_file_item.name}' 처리 실패: {current_file_error_message}")
                        st.session_state.failed_files_info.append({'name': uploaded_file_item.name, 'error': current_file_error_message})
            
            # 모든 파일 처리 후 다운로드 데이터 생성 (전체 누적 데이터 사용)
            if st.session_state.all_accumulated_rows:
                st.session_state.download_data = generate_excel_for_download(st.session_state.all_accumulated_rows)
                total_processed = len(st.session_state.processed_files)
                st.sidebar.success(f"총 {total_processed}개 파일 처리 완료.")
                if st.session_state.failed_files_info:
                    st.sidebar.error(f"이번 시도에서 {len(st.session_state.failed_files_info)}개 파일 처리 실패.")
            else:
                st.sidebar.warning("처리할 수 있는 데이터가 없습니다.")

    # --- 메인 영역 ---
    st.markdown("<h1 style='text-align: center;'>📊 엑셀 데이터 종합 처리기</h1>", unsafe_allow_html=True)
    st.markdown("---")

    if st.session_state.failed_files_info: # 실패한 파일 정보 표시
        with st.expander("파일 처리 실패 정보 보기", expanded=False):
            for failed_file in st.session_state.failed_files_info:
                st.error(f"파일: {failed_file['name']}, 오류: {failed_file['error']}")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📜 종합 결과 미리보기")
        if st.session_state.all_accumulated_rows:
            headers = ["위치", "계측기명", "현재값"]
            df_display_combined = pd.DataFrame(st.session_state.all_accumulated_rows, columns=headers)
            st.dataframe(df_display_combined, height=400, use_container_width=True)
        elif not uploaded_file_list: #초기 상태
             st.info("왼쪽 사이드바에서 처리할 Excel 파일들을 업로드해주세요.")
        else: # 파일은 업로드되었으나 처리된 데이터가 없는 경우
             st.warning("처리된 데이터가 없습니다. 업로드된 파일의 내용을 확인하거나, 실패 정보를 확인해주세요.")

    with col2:
        st.subheader("💾 종합 결과 다운로드")
        if st.session_state.download_data:
            # 날짜를 파일명에 사용
            file_date = st.session_state.last_date if st.session_state.last_date else "결과"
            download_filename = f"종합_결과_{file_date}.xlsx"
            
            st.download_button(
                label="종합 결과 Excel 파일 다운로드",
                data=st.session_state.download_data,
                file_name=download_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("다운로드할 종합 데이터가 없습니다.")
    
    st.markdown("---")
    st.subheader("💡 사용 방법")
    st.markdown("""
    1.  왼쪽 사이드바에서 **하나 이상의 Excel 파일**을 업로드하세요.
    2.  파일이 업로드되면 자동으로 각 파일에 대한 데이터 처리가 시작됩니다.
    3.  처리가 완료되면, 종합 결과 미리보기 영역에서 모든 파일의 처리 결과가 순서대로 쌓인 내용을 확인할 수 있습니다.
    4.  (각 파일의 [시트제목+첫 행 값]이 헤더로, 해당 파일의 마지막 행 값이 데이터로 표시됩니다.)
    5.  다운로드 버튼을 사용하여, 이 종합 결과를 단일 Excel 시트로 받을 수 있습니다.
    6.  개별 파일 처리 실패 시, 실패 정보가 별도로 표시됩니다.
    """)

if __name__ == "__main__":
    main() 