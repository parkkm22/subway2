import streamlit as st
import pandas as pd
from datetime import datetime
import os
import google.generativeai as genai
import io
import PyPDF2
import re
import pdfplumber
from pdf2image import convert_from_bytes
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 발파데이터 추출 프롬프트 (간결 버전)
BLAST_EXTRACTION_PROMPT = '''
# Instruction
다음 두 개의 TSV 형식 문자열 (첫 번째는 발파작업일지 TSV, 두 번째는 계측일지 TSV)을 입력으로 받아, 지정된 규칙에 따라 데이터를 추출하고 병합하여 최종 TSV(UTF-8) 형식으로 출력하세요.

## 입력 1: 발파작업일지_TSV
- 내용: 발파작업일지 PDF에서 추출된 데이터입니다.
- 컬럼 예시: 발파일자, 발파시간, 지발당장약량(최소, kg), 지발당장약량(최대, kg), 폭약사용량(kg), 비고
- 실제 데이터 예시 (아래 형식으로 제공됨):
발파일자\\t발파시간\\t지발당장약량(최소, kg)\\t지발당장약량(최대, kg)\\t폭약사용량(kg)\\t비고
2023-07-27\\t08:05\\t0.5\\t0.9\\t73\\tPLA-2
2023-07-27\\t13:47\\t0.4\\t0.8\\t77\\tPD-2
(참고: 실제 입력되는 컬럼명과 값, 날짜/시간 형식은 원본 PDF의 것을 따릅니다.)

## 입력 2: 계측일지_TSV
- 내용: 발파진동소음 계측일지 PDF에서 추출된 데이터입니다. \\\'측정위치\\\'는 이미 각 행에 포함되어 있습니다.
- 컬럼 예시: Date/Time, Peak Particle Vel (X_Axis) (mm/sec), Peak Particle Vel (Y_Axis) (mm/sec), Peak Particle Vel (Z_Axis) (mm/sec), LMax (Sound) (dBA), 측정위치
- 실제 데이터 예시 (아래 형식으로 제공됨):
Date/Time\\tPeak Particle Vel (X_Axis) (mm/sec)\\tPeak Particle Vel (Y_Axis) (mm/sec)\\tPeak Particle Vel (Z_Axis) (mm/sec)\\tLMax (Sound) (dBA)\\t측정위치
2023/07/27 1:47:00 PM\\t0.71\\t0.36\\t0.71\\t61.23\\t양말집
2023/07/27 1:47:00 PM\\t0.87\\t0.56\\t0.87\\t53.29\\t티스테이션
(참고: 실제 입력되는 컬럼명, 값, 날짜/시간 형식은 원본 PDF의 것을 따릅니다.)

# Output
1. 형식: TSV(UTF-8)
2. 정렬 기준: "발파시간" 오름차순, 그 다음 "계측위치" 오름차순 (필요시)
3. 최종 헤더(고정열): 발파일자, 발파시간, 지발당장약량(최소, kg), 지발당장약량(최대, kg), 폭약사용량(kg), 발파진동(cm/sec), 발파소음(dB(A)), 계측위치, 비고
4. **중요 출력 규칙**: 오직 최종 TSV 데이터만 출력해야 하며, 다른 설명, 분석 내용, 주석, 마크다운 형식은 절대 포함하지 마십시오.

# 추출, 변환 및 병합 규칙
(이하 규칙은 이전과 동일)
1.  **기준**: \\\'발파작업일지_TSV\\\'의 각 행을 기준으로, \\\'계측일지_TSV\\\'와 데이터를 병합합니다.
2.  **발파일자**: \\\'발파작업일지_TSV\\\'의 날짜 정보를 YYYY-MM-DD 형식으로 사용합니다.
3.  **발파시간 (매칭 기준)**: \\\'발파작업일지_TSV\\\'의 시간 정보를 HH:MM (24시간제) 형식으로 변환하여 매칭 기준으로 사용합니다.
4.  **지발당장약량(최소, kg)**, **지발당장약량(최대, kg)**, **폭약사용량(kg)**, **비고**:
    - \\\'발파작업일지_TSV\\\'에서 해당 발파시간의 값을 가져옵니다.
5.  **계측일지 시간 변환**: \\\'계측일지_TSV\\\'의 시간 정보를 HH:MM (24시간제) 형식으로 변환합니다.
6.  **데이터 매칭**: \\\'발파작업일지_TSV\\\'의 (변환된) \\\'발파시간\\\'과 \\\'계측일지_TSV\\\'의 (변환된) \\\'시간\\\'을 비교하여 일치하는 모든 행을 찾습니다.
7.  **발파진동(cm/sec)**:
    1.  시간이 매칭된 \\\'계측일지_TSV\\\' 행에서 X, Y, Z축 진동 값 중 **최댓값**을 선택 (단위: mm/sec).
    2.  선택된 값을 10으로 나눠 cm/sec로 변환 후 소수점 셋째 자리까지 표시.
    3.  매칭되는 계측 데이터가 없으면 \\\"-\\\"로 표시.
8.  **발파소음(dB(A))**: 시간이 매칭된 \\\'계측일지_TSV\\\' 행의 소음 값을 소수점 셋째 자리까지 표시. 없으면 \\\"-\\\".
9.  **계측위치**: 시간이 매칭된 \\\'계측일지_TSV\\\' 행의 \\\'측정위치\\\' 값을 사용. 없으면 \\\"-\\\".
10. **다중 계측 데이터 처리 (중요)**:
    - 하나의 \\\'발파작업일지_TSV\\\' 행(특정 발파시간)에 대해, \\\'계측일지_TSV\\\'에서 여러 개의 매칭되는 행이 있을 수 있습니다 (보통 계측위치가 다른 경우).
    - 이 경우, 최종 출력 테이블에 각 매칭 건당 하나의 행을 생성합니다.
    - 첫 번째 생성된 행(해당 발파시간의 첫 번째 계측 데이터)에만 **지발당장약량(최소, kg)**, **지발당장약량(최대, kg)**, **폭약사용량(kg)** 값을 표시합니다.
    - 동일 발파시간의 두 번째 이후 매칭 행에는 이 세 가지 값을 \\\"-\\\"로 표시합니다.
    - **발파일자, 발파시간, 비고**는 해당 발파시간의 모든 매칭 행에 동일하게 반복 표시합니다.
    - **발파진동(cm/sec), 발파소음(dB(A)), 계측위치**는 각 매칭된 계측일지 행의 값을 따릅니다.
11. **숫자 포맷 및 값 부재**: 모든 수치는 요구되는 소수점 자리까지 표기하고, 값이 없거나 매칭되지 않으면 \\\"-\\\"로 표시합니다. 모든 텍스트 값은 그대로 유지합니다.

# 최종 출력 예시 (아래 예시는 최종 출력 형식을 보여주기 위함이며, 실제 내용은 입력 데이터에 따라 동적으로 생성되어야 합니다.)
발파일자\\t발파시간\\t지발당장약량(최소, kg)\\t지발당장약량(최대, kg)\\t폭약사용량(kg)\\t발파진동(cm/sec)\\t발파소음(dB(A))\\t계측위치\\t비고
[날짜]\\t[시간1]\\t[값]\\t[값]\\t[값]\\t-\\t-\\t-\\t[비고1]
[날짜]\\t[시간2]\\t[값]\\t[값]\\t[값]\\t[값]\\t[값]\\t[위치1]\\t[비고2]
[날짜]\\t[시간2]\\t-\\t-\\t-\\t[값]\\t[값]\\t[위치2]\\t[비고2]
[날짜]\\t[시간3]\\t[값]\\t[값]\\t[값]\\t[값]\\t[값]\\t[위치3]\\t[비고3]
[날짜]\\t[시간3]\\t-\\t-\\t-\\t[값]\\t[값]\\t[위치4]\\t[비고3]
[날짜]\\t[시간4]\\t[값]\\t[값]\\t[값]\\t-\\t-\\t-\\t[비고4]

**매우 중요**: 다른 어떤 텍스트, 설명, 분석, 주석, 마크다운도 없이, 위에 제시된 \\\'최종 출력 예시\\\'와 정확히 동일한 형식의 TSV 데이터만 출력해야 합니다. 이 지침을 반드시 엄격하게 준수하십시오.
'''

# Gemini API 키 설정
GENAI_API_KEY = "AIzaSyD69-wKYfZSID327fczrkx-JveJdGYIUIk" # 이 부분은 실제 키로 대체되어야 합니다.
genai.configure(api_key=GENAI_API_KEY)

# Gemini 모델 객체 생성
GEMINI_MODEL = genai.GenerativeModel("models/gemini-2.5-flash-preview-05-20")

DEFAULT_PROMPT = """
# INSTRUCTIONS
1. 기상청 서울 지역 관측 자료를 기반으로 "날씨정보" 테이블을 TSV(UTF-8) 형식의 별도 코드블록으로 생성
2. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비") 각각을 TSV(UTF-8) 형식의 별도 코드블록으로 차례대로 출력하며 아래의 조건을 철저히 준수할 것

# OUTPUT : 테이블(총 5개)  
## 1. 날씨정보 테이블
1. 고정 열 : "구분", "값"
2. 고정 행 : "최고온도", "최저온도", "강수량"
3. 추출데이터 : 서울(유) 오늘 날씨 예보 (최신 업데이트)
4. 주의사항 
- 서울 지역(영등포구 우선)의 최고 기온, 최저 기온, 강수량의 단일값 추출
- 데이터는 최신 업데이트된 기상청 정보를 기반으로 제공
- "값"만 숫자로 추출할 것 (예: 20.0 °C에서 "20.0" 추출)

## 2. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 33행) - 아래 순서와 명칭을 그대로  
- "1. 본선터널 (1구간, 대림-신풍) 굴착"  
- "1. 본선터널 (1구간, 대림-신풍) 라이닝" 
- "2. 신풍정거장 - 1)정거장 라이닝"
- "2. 신풍정거장 - 1)정거장 미들 슬라브"
- "2. 신풍정거장 – 2)주출입구 수직구 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 정거장 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 환승통로 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (2)PCC 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (3)PCD 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (4)PHA 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - 수직구 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - PHB 라이닝"
- "2. 신풍정거장 - 4)외부출입구 출입구(#3) 굴착" 
- "2. 신풍정거장 - 4)외부출입구 출입구(#2) 굴착"
- "2. 신풍정거장 - 4)외부출입구 출입구(#1) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 라이닝"  
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 라이닝"  
- "3. 신풍 환승통로 - 2)개착 BOX 보라매 방면 구조물"  
- "3. 신풍 환승통로 - 2)개착 BOX 대림 방면 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 터널 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 미들 슬라브" 
- "5. 도림사거리정거장 - 2)출입구#1 수직구 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCA 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCC 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PHA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 수직구 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCC 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PHB 라이닝"  
3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예: 945.3m / 1,116m 에서 "945.3" 추출)
 
## 3. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 14행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍-도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"  
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\n' 문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, '-'기호는 생략함

## 4. 인원 / 장비 테이블  
1. 고정 열 (총 15열) - 열 순서는 아래와 같음
- "구분" 
- "1. 본선터널 (1구간, 대림~신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍~도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"    

2. 고정 행(인원 테이블 – 총 36행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"직영반장", "연수생", "장비운전원", "전기주임", "화약주임", "터널공", "목공", "철근공", "라이닝폼공", "오폐수처리공", "카리프트공", "BP공", "가시설공", "설치공/해체공", "동바리공", "신호수", "부단수공", "슬러리월공", "CIP공", "미장공", "시설물공", "경계석공", "조경공", "배관공", "도색공", "방수공", "장비/작업지킴이", "보통인부", "포장공", "용접공", "타설공", "보링공/앙카공", "비계공", "도장공", "석면공", "주입공/그라우팅공"

3. 고정 행 (장비 테이블 – 총 46행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):

"B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)", "덤프트럭(5T)", "덤프트럭(15T)", "덤프트럭(25T)", "앵글크레인(100T)", "앵글크레인(80T)", "앵글크레인(35T)", "앵글크레인(25T)", "카고크레인(25T)", "카고크레인(5T)", "콤프", "점보드릴", "페이로더", "숏트머신", "차징카", "살수차", "하이드로크레인", "믹서트럭", "화물차(5T)", "펌프카", "스카이", "콘크리트피니셔", "전주오거", "로더(바브켓)", "유제살포기(비우다)", "지게차", "싸인카", "BC커터기", "바이브로해머", "롤러(2.5T)", "롤러(1T)", "롤러(0.7T)", "몰리", "항타기", "크레인", "콤비로라", "공압드릴", "유압드릴", "기타"

## 5. Parsing Rules 
1. 시공현황: "누계/설계" → **앞 값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑    
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤, PHA → ⑥, 특별피난 → ⑦, 외부출입구 → ⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" → "B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.
    
## 6. QA-CHECKLIST(자동 검증 후 변환로그 출력)
- **구조**: 테이블/행·열 개수·순서 일치?    
- **빈셀 0** 규칙 준수?    
- **데이터**: 시공현황 숫자, 인원·장비 정수?    
- **매핑**: 변환 누락 여부?    
- **누락/중복**: 원문 수량과 100 % 일치?  
- **변환로그**: 매핑·정규화 과정에서 바뀐 원/결과 목록(있을 때만 출력)
"""

def convert_pdf_to_images(pdf_file):
    """PDF를 이미지로 변환하는 함수"""
    try:
        # PDF 파일을 이미지로 변환 (300 DPI로 고품질)
        images = convert_from_bytes(pdf_file.read(), dpi=300)
        pdf_file.seek(0)  # 파일 포인터 리셋
        return images
    except Exception:
        return []

def extract_table_from_images_with_llm(images, file_type="계측일지"):
    """이미지에서 LLM을 사용하여 표 데이터를 추출하는 함수"""
    if not images:
        return None
    
    try:
        # 표 추출 전용 프롬프트
        if file_type == "계측일지":
            image_prompt = '''
이 이미지에서 발파진동·소음 계측 데이터 표를 찾아 모든 측정값을 추출해주세요.

**중요 지시사항**:
1. Event List, 측정결과, 계측데이터 등의 표에서 데이터 추출
2. 모든 시간대의 데이터를 빠뜨리지 말고 추출
3. AM/PM 시간을 24시간 형식으로 변환 (예: 1:47 PM → 13:47, 8:05 AM → 08:05)

**추출할 데이터**:
- 시간 (Time) - HH:MM 24시간 형식으로 변환
- Peak Particle Velocity X축 (mm/sec)
- Peak Particle Velocity Y축 (mm/sec) 
- Peak Particle Velocity Z축 (mm/sec)
- LMax Sound Level (dBA)
- 측정위치/계측위치

**출력 형식 (TSV)**:
시간	X_Axis	Y_Axis	Z_Axis	Sound	측정위치
08:05	0.710	0.650	0.580	61.2	도림사거리
13:47	0.880	0.750	0.620	53.2	도림사거리
16:10	0.920	0.680	0.590	-	도림사거리
16:40	0.850	0.790	0.650	61.2	도림사거리

**주의사항**:
- 각 측정 지점마다 별도 행으로 출력
- 값이 없거나 측정되지 않은 경우 "-" 표시
- 헤더는 위 형식 그대로 사용
- TSV 형식 (탭으로 구분)
- 소수점 3자리까지 표시
'''
        else:  # 발파작업일지
            image_prompt = '''
이 이미지에서 발파 작업 관련 표의 데이터를 추출해서 TSV 형식으로 변환해주세요.

추출할 컬럼:
- 발파일자
- 발파시간
- 지발당장약량(최소, kg)
- 지발당장약량(최대, kg)
- 폭약사용량(kg)
- 비고

출력 형식:
- TSV (탭으로 구분)
- 헤더 포함
- 숫자는 소수점 3자리까지
- 날짜는 YYYY-MM-DD 형식
- 시간은 HH:MM 형식
'''
        
        # 첫 번째 이미지부터 순차적으로 시도
        best_result = None
        for i, image in enumerate(images):
            # 이미지를 Gemini에 전송
            response = GEMINI_MODEL.generate_content([image_prompt, image])
            
            if response.text and '\t' in response.text:
                # 더 많은 데이터를 포함한 결과를 우선선택
                if not best_result or len(response.text.split('\n')) > len(best_result.split('\n')):
                    best_result = response.text
        
        return best_result
        
    except Exception:
        return None

def extract_target_table(pdf_file):
    """PDF에서 특정 테이블을 추출하는 함수 (텍스트 + 이미지 방식)"""
    
    # 1차 시도: 텍스트 기반 추출
    found_tables = []
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page_num, page in enumerate(pdf.pages):
                for table_num, table in enumerate(page.extract_tables()):
                    if table and len(table) > 1:
                        # 헤더가 None이 아닌 경우에만 처리
                        clean_headers = [str(col).strip() if col is not None else f"Col_{i}" for i, col in enumerate(table[0])]
                        
                        # 데이터 행 정리
                        clean_data = []
                        for row in table[1:]:
                            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                            clean_data.append(clean_row)
                        
                        df = pd.DataFrame(clean_data, columns=clean_headers)
                        
                        # 발파 관련 키워드가 포함된 테이블 찾기 (조건 완화)
                        column_text = ' '.join([str(col) for col in df.columns if col])
                        data_text = ' '.join([str(cell) for row in clean_data[:3] for cell in row])  # 첫 3행만 확인
                        all_text = (column_text + ' ' + data_text).lower()
                        
                        # 더 넓은 범위의 키워드로 검색
                        blast_keywords = ['발파', '시간', '장약', '폭약', '비고', '일자', '날짜', 'time', 'blast', 'charge']
                        
                        if any(keyword in all_text for keyword in blast_keywords):
                            found_tables.append({
                                'page': page_num + 1,
                                'table': table_num + 1,
                                'dataframe': df,
                                'columns': clean_headers,
                                'text_sample': all_text[:200]
                            })
    except Exception as e:
        print(f"텍스트 기반 추출 오류: {e}")
    
    # 가장 적합한 테이블 선택 (더 많은 컬럼을 가진 것 우선)
    if found_tables:
        best_table = max(found_tables, key=lambda x: len(x['dataframe'].columns))
        return best_table['dataframe']
    
    # 2차 시도: 이미지 기반 추출
    pdf_file.seek(0)  # 파일 포인터 리셋
    images = convert_pdf_to_images(pdf_file)
    
    if images:
        table_text = extract_table_from_images_with_llm(images, "작업일지")
        if table_text:
            try:
                # TSV 텍스트를 DataFrame으로 변환
                tsv_data = extract_tsv_from_response(table_text)
                tsv_data = fix_tsv_field_count(tsv_data)
                df = pd.read_csv(io.StringIO(tsv_data), sep='\t', encoding='utf-8')
                return df
            except Exception:
                pass
    
    return None

def extract_file_content(file):
    """파일에서 텍스트 내용을 추출하는 함수 (Gemini 직접 PDF 처리)

    각 PDF에서 표 형태의 데이터를 최대한 원본에 가깝게 TSV로 추출하는 데 집중합니다.
    복잡한 데이터 변환이나 파일 간 병합은 메인 프롬프트에서 처리합니다.
    """
    if file.name.endswith('.pdf'):
        try:
            file_content = file.read()
            file.seek(0)
            if not file_content.startswith(b'%PDF'):
                st.error("❌ 업로드된 파일이 유효한 PDF 형식이 아닙니다.")
                return None
        except Exception as e:
            st.error(f"❌ 파일 읽기 실패: {str(e)}")
            return None

        filename_lower = file.name.lower()
        is_measurement_file = any(keyword in filename_lower for keyword in 
                                 ["계측", "진동", "소음", "measurement", "vibration", "noise"])
        is_blast_log_file = any(keyword in filename_lower for keyword in 
                                 ["발파", "작업", "일지", "blast", "work", "log"])

        pdf_prompt = ""
        if is_measurement_file:
            pdf_prompt = '''
이 PDF 파일은 '발파진동소음 계측일지'입니다. 다음 지침에 따라 데이터를 TSV 형식으로 추출해주세요.

**중요 지시사항**:
1.  PDF 페이지를 순서대로 읽으면서 "계측 : [위치명]" 또는 "Location : [Location Name]" 패턴을 찾습니다.
2.  해당 패턴 바로 다음에 나오는 데이터 테이블에서 다음 원본 컬럼들의 모든 행을 추출합니다:
    - "Date/Time" (또는 유사한 날짜/시간 컬럼)
    - "Peak Particle Vel (X_Axis) (mm/sec)" (또는 X축 진동 관련 컬럼)
    - "Peak Particle Vel (Y_Axis) (mm/sec)" (또는 Y축 진동 관련 컬럼)
    - "Peak Particle Vel (Z_Axis) (mm/sec)" (또는 Z축 진동 관련 컬럼)
    - "LMax (Sound) (dBA)" (또는 소음 관련 컬럼)
3.  추출된 데이터에 '측정위치' 컬럼을 추가하고, 1번에서 찾은 [위치명]을 모든 행에 반복하여 입력합니다.
4.  날짜/시간은 원본 그대로 (예: "7/27/2023 1:47:39 PM") 추출합니다. 변환은 나중에 합니다.

**출력 TSV 헤더 및 예시 (실제 추출되는 컬럼명은 원본 PDF를 따름)**:
Date/Time	Peak Particle Vel (X_Axis) (mm/sec)	Peak Particle Vel (Y_Axis) (mm/sec)	Peak Particle Vel (Z_Axis) (mm/sec)	LMax (Sound) (dBA)	측정위치
7/27/2023 1:47:39 PM	0.62	0.36	0.71	61.23	양말집
7/27/2023 4:10:12 PM	0.29	0.28	0.59	58.15	양말집
7/27/2023 4:10:08 PM	0.26	0.44	0.88	47.96	티스테이션
7/27/2023 1:47:34 PM	0.60	0.56	0.87	53.29	티스테이션

값이 없거나 해당 컬럼이 존재하지 않으면 빈 문자열 또는 "-"로 표시하세요. 다른 설명 없이 TSV 데이터만 출력하세요.
'''
        elif is_blast_log_file:
            pdf_prompt = '''
이 PDF 파일은 '발파작업일지'입니다. 다음 지침에 따라 주요 데이터를 TSV 형식으로 추출해주세요.

**추출 대상 컬럼 (PDF에 있는 실제 컬럼명 기준, 아래는 표준 예시)**:
- 발파일자 (또는 Date, 날짜 등)
- 발파시간 (또는 Time, 시간 등)
- 지발당장약량(최소, kg) (또는 Min Charge per Hole 등)
- 지발당장약량(최대, kg) (또는 Max Charge per Hole 등)
- 폭약사용량(kg) (또는 Total Explosives 등)
- 비고 (또는 Remarks, 기타 등)

**출력 형식 (TSV)**:
- 헤더는 PDF의 실제 컬럼명을 따르세요.
- 모든 관련 데이터를 추출하세요.
- 날짜와 시간은 원본 형식 그대로 추출하세요.

**예시 (실제 컬럼명과 데이터는 PDF 내용에 따라 다름)**:
발파일자	발파시간	지발당장약량(최소, kg)	지발당장약량(최대, kg)	폭약사용량(kg)	비고
2023-10-26	13:47	0.050	0.100	150.000	암반 발파 (굴착 작업)
2023-10-26	16:10	0.060	0.120	180.000	2차 발파 (터널 확장)

다른 설명 없이 TSV 데이터만 출력하세요.
'''
        else:
            # 파일 유형을 특정할 수 없는 경우, 일반적인 표 추출 시도 (선택적)
            st.warning("⚠️ 파일 유형을 특정할 수 없어 일반 표 추출을 시도합니다. 결과가 정확하지 않을 수 있습니다.")
            pdf_prompt = "이 PDF에서 가장 중요해 보이는 표를 찾아 TSV 형식으로 추출해주세요. 다른 설명은 생략하고 TSV 데이터만 출력하세요."

        if not pdf_prompt:
             st.error("❌ PDF 파일 유형을 인식할 수 없습니다. (발파작업일지 또는 계측일지)")
             return None

        try:
            file.seek(0)
            uploaded_file = genai.upload_file(file, mime_type="application/pdf")
            response = GEMINI_MODEL.generate_content([pdf_prompt, uploaded_file])
            genai.delete_file(uploaded_file.name)

            if response.text:
                # 응답 앞뒤의 마크다운 코드 블록 제거 및 공백 제거
                cleaned_response = response.text.strip()
                if cleaned_response.startswith("```tsv"):
                    cleaned_response = cleaned_response[len("```tsv"):].strip()
                elif cleaned_response.startswith("```"):
                    cleaned_response = cleaned_response[len("```"):].strip()
                if cleaned_response.endswith("```"):
                    cleaned_response = cleaned_response[:-len("```")].strip()
                return cleaned_response
            else:
                st.error(f"❌ {file.name}에서 AI가 내용을 추출하지 못했습니다.")
                return None

        except Exception as e:
            st.error(f"❌ {file.name} 처리 중 AI 오류 발생: {str(e)}")
            # 백업 로직 (PyPDF2, pdfplumber 등)은 여기서 제거하고, 필요시 메인 로직에서 재시도 고려
            return None
        
    elif file.name.endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(file, engine='openpyxl')
            df = df.fillna('-') 
            # 엑셀은 비교적 정형화되어 있으므로, 여기서 간단한 문자열 변환 후 TSV로 반환
            return df.to_csv(sep='\t', index=False, encoding='utf-8')
        except Exception as e:
            st.error(f"❌ 엑셀 데이터 추출 실패: {str(e)}")
            return None
    elif file.name.endswith('.docx'):
        st.warning("⚠️ 워드 파일 처리 기능은 아직 구현되지 않았습니다.")
        return None # 또는 "워드 파일 처리 기능은 아직 구현되지 않았습니다."
    
    st.error(f"❌ 지원하지 않는 파일 형식입니다: {file.name}")
    return None

def parse_tsv_to_dataframe(tsv_content):
    """TSV 형식의 문자열을 DataFrame으로 변환
    
    Args:
        tsv_content (str): TSV 형식의 문자열
        
    Returns:
        pd.DataFrame: 변환된 DataFrame 또는 None
    """
    try:
        # 빈 줄 제거 및 공백 문자 처리
        cleaned_content = '\n'.join(line.strip() for line in tsv_content.split('\n') if line.strip())
        
        # TSV 문자열을 DataFrame으로 변환
        df = pd.read_csv(io.StringIO(cleaned_content), sep='\t', encoding='utf-8')
        
        # 컬럼명에서 불필요한 공백 제거
        df.columns = df.columns.str.strip()
        
        # None 값을 빈 문자열로 변환
        df = df.fillna('')
        
        return df
    except Exception as e:
        st.error(f"TSV 파싱 중 오류 발생: {str(e)}")
        return None

def extract_tsv_from_response(response_text):
    """LLM 응답에서 TSV 데이터를 추출하는 함수 (강화된 버전)
    
    Args:
        response_text (str): LLM의 응답 텍스트
        
    Returns:
        str: 추출된 TSV 데이터 (앞뒤 공백 및 코드블록 마커 제거됨)
             추출 실패 시 원본 텍스트의 주요 부분 또는 빈 문자열 반환 시도.
    """
    if not response_text:
        return ""

    lines = response_text.strip().split('\n')
    cleaned_lines = []

    # 1. 코드 블록 마커 및 언어 지정자 제거 시도
    temp_lines = []
    in_code_block = False
    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith("```") and not in_code_block:
            # 코드 블록 시작 (```tsv, ```python 등) 또는 ``` 단독
            in_code_block = True
            # ``` 다음의 언어 지정자 제거 (예: ```tsv -> 공백)
            content_after_ticks = stripped_line[3:].strip()
            if content_after_ticks and not '\t' in content_after_ticks: # 언어 지정자일 가능성
                continue # 언어 지정자 라인 자체는 추가 안 함
            elif '\t' in content_after_ticks: # ``` 뒤에 바로 TSV 데이터가 온 경우
                 temp_lines.append(content_after_ticks)
            continue
        elif stripped_line.startswith("```") and in_code_block:
            # 코드 블록 끝
            in_code_block = False
            continue
        
        if in_code_block:
            temp_lines.append(line) # 코드 블록 내부는 그대로 유지
        else:
            # 코드 블록 바깥의 내용은 일단 모두 추가 (나중에 TSV인지 검사)
            temp_lines.append(line)
    
    lines = temp_lines # 코드 블록 처리된 라인들로 교체

    # 2. TSV 데이터 영역 찾기 (헤더 기반)
    # BLAST_EXTRACTION_PROMPT의 최종 헤더와 유사한 패턴을 찾음
    expected_headers = ["발파일자", "발파시간", "지발당장약량(최소, kg)", "지발당장약량(최대, kg)", 
                        "폭약사용량(kg)", "발파진동(cm/sec)", "발파소음(dB(A))", "계측위치", "비고"]
    
    tsv_start_index = -1
    best_match_count = 0

    for i, line in enumerate(lines):
        # 현재 라인이 헤더일 가능성이 있는지 검사 (최소 3개 이상 헤더 키워드 포함)
        current_line_headers = [header for header in expected_headers if header in line]
        if len(current_line_headers) >= 3 and '\t' in line:
            # 더 많은 헤더 키워드를 포함하는 라인을 우선적으로 헤더로 간주
            if len(current_line_headers) > best_match_count:
                best_match_count = len(current_line_headers)
                tsv_start_index = i
    
    if tsv_start_index != -1:
        # 헤더부터 시작하여 유효한 TSV 라인들만 추출
        for i in range(tsv_start_index, len(lines)):
            line_to_check = lines[i].strip()
            # 최소 1개 이상의 탭이 있고, 데이터로 보이는 라인만 포함
            # (너무 짧거나, 특정 설명조의 문장부호로 시작하는 등 제외 가능)
            if '\t' in line_to_check and len(line_to_check.split('\t')) > 2: # 최소 3컬럼 이상
                cleaned_lines.append(line_to_check)
            elif cleaned_lines: # 이미 TSV 데이터 수집 중이었다면, 여기서 중단
                break
    
    # 3. 헤더 기반으로 못 찾았을 경우, 전체 라인 중 탭 포함 라인만 선택 (최후의 수단)
    if not cleaned_lines:
        for line in lines:
            if '\t' in line.strip():
                cleaned_lines.append(line.strip())

    if not cleaned_lines:
        # 그래도 못 찾으면, 원본에서 앞부분 일부라도 반환 시도 (디버깅용)
        # st.warning("TSV 데이터를 정확히 추출하지 못했습니다. AI 응답 일부를 확인하세요.")
        return "\n".join(lines[:15]).strip() # 최대 15줄

    return "\n".join(cleaned_lines)

# TSV 필드 개수 자동 보정 함수 추가
def fix_tsv_field_count(tsv_str):
    """TSV 데이터의 필드 수를 헤더와 맞춰 보정하는 함수 (개선된 버전)"""
    lines = tsv_str.strip().split('\n')
    if not lines:
        return tsv_str
    
    header = lines[0]
    n_fields = header.count('\t') + 1
    fixed_lines = [header]
    
    for i, line in enumerate(lines[1:], 2):
        fields = line.split('\t')
        
        # 필드가 부족한 경우 빈 문자열로 채움
        if len(fields) < n_fields:
            fields += [''] * (n_fields - len(fields))
        # 필드가 많은 경우 마지막 필드에 합침
        elif len(fields) > n_fields:
            # 텍스트 데이터에 탭이 포함된 경우를 고려하여 마지막 필드에 결합
            extra_fields = fields[n_fields-1:]
            fields = fields[:n_fields-1] + [' '.join(extra_fields)]
        
        fixed_lines.append('\t'.join(fields))
    
    return '\n'.join(fixed_lines)

def validate_and_clean_tsv(tsv_str):
    """TSV 데이터를 검증하고 정제하는 함수"""
    if not tsv_str or not tsv_str.strip():
        return ""
    
    lines = tsv_str.strip().split('\n')
    if not lines:
        return ""
    
    # 빈 라인 제거
    lines = [line for line in lines if line.strip()]
    
    if not lines:
        return ""
    
    # 헤더 검증
    header = lines[0]
    if '\t' not in header:
        # 헤더에 탭이 없으면 TSV가 아닐 가능성
        return ""
    
    n_fields = header.count('\t') + 1
    cleaned_lines = [header]
    
    # 각 라인 검증 및 정제
    for i, line in enumerate(lines[1:], 2):
        # 라인이 너무 비정상적으로 길면 건너뛰기
        if len(line) > 10000:  # 10,000자 이상인 라인은 건너뛰기
            continue
            
        fields = line.split('\t')
        
        # 필드 수 조정
        if len(fields) < n_fields:
            fields += [''] * (n_fields - len(fields))
        elif len(fields) > n_fields:
            # 초과 필드를 마지막 필드에 합치되, 합리적인 길이로 제한
            extra_text = ' '.join(fields[n_fields-1:])
            if len(extra_text) > 1000:  # 너무 긴 텍스트는 잘라내기
                extra_text = extra_text[:1000] + "..."
            fields = fields[:n_fields-1] + [extra_text]
        
        cleaned_lines.append('\t'.join(fields))
    
    return '\n'.join(cleaned_lines)

def remove_tsv_label(tsv_str):
    lines = tsv_str.strip().split('\n')
    if lines and lines[0].strip().lower() == "tsv":
        return '\n'.join(lines[1:])
    return tsv_str

def insert_blast_data_to_excel(blast_df, template_file, start_row=2, start_col=1):
    """
    발파데이터를 샘플 엑셀 파일의 지정된 위치에 입력하는 함수
    
    Args:
        blast_df (pd.DataFrame): 발파데이터 DataFrame
        template_file: 업로드된 샘플 엑셀 파일
        start_row (int): 데이터를 입력할 시작 행 (1-based)
        start_col (int): 데이터를 입력할 시작 열 (1-based)
    
    Returns:
        bytes: 수정된 엑셀 파일의 바이트 데이터
    """
    try:
        # 엑셀 파일 로드
        wb = load_workbook(template_file)
        ws = wb.active  # 첫 번째 워크시트 사용
        
        # 헤더 입력 (선택사항) - B119에는 헤더 불필요하므로 주석 처리
        # for col_idx, column_name in enumerate(blast_df.columns):
        #     ws.cell(row=start_row-1, column=start_col+col_idx, value=column_name)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(blast_df.itertuples(index=False)):
            for col_idx, cell_value in enumerate(row_data):
                ws.cell(row=start_row+row_idx, column=start_col+col_idx, value=cell_value)
        
        # 바이트 스트림으로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ 엑셀 파일 처리 중 오류 발생: {str(e)}")
        return None

def insert_blast_data_to_excel_ae160(blast_df, template_file, start_row=160, start_col=31):
    """
    발파데이터를 샘플 엑셀 파일의 AE160 위치에 입력하는 함수
    
    Args:
        blast_df (pd.DataFrame): 발파데이터 DataFrame
        template_file: 업로드된 샘플 엑셀 파일
        start_row (int): 데이터를 입력할 시작 행 (AE160의 160)
        start_col (int): 데이터를 입력할 시작 열 (AE열 = 31)
    
    Returns:
        bytes: 수정된 엑셀 파일의 바이트 데이터
    """
    try:
        # 엑셀 파일 로드
        wb = load_workbook(template_file)
        ws = wb.active  # 첫 번째 워크시트 사용
        
        # 헤더 입력하지 않음 (AE160에는 헤더 불필요)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(blast_df.itertuples(index=False)):
            for col_idx, cell_value in enumerate(row_data):
                ws.cell(row=start_row+row_idx, column=start_col+col_idx, value=cell_value)
        
        # 바이트 스트림으로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ 엑셀 파일 처리 중 오류 발생: {str(e)}")
        return None

def insert_five_tables_to_excel(tables_data, template_file, table_positions=None):
    """
    5개 테이블(날씨정보, 시공현황, 작업내용, 인원, 장비)을 엑셀 파일의 지정된 위치에 입력하는 함수
    
    Args:
        tables_data (list): 5개 테이블의 DataFrame 리스트
        template_file: 업로드된 샘플 엑셀 파일
        table_positions (dict): 각 테이블의 위치 정보 (기본값 사용)
    
    Returns:
        bytes: 수정된 엑셀 파일의 바이트 데이터
    """
    try:
        # 엑셀 파일 로드
        wb = load_workbook(template_file)
        ws = wb.active  # 첫 번째 워크시트 사용
        
        # 기본 테이블 위치 설정 (사용자 요청에 따라 수정)
        if table_positions is None:
            table_positions = {
                "날씨정보": {"row": 5, "col": 30},    # AD5
                "시공현황": {"row": 13, "col": 30},   # AD13
                "작업내용": {"row": 48, "col": 30},   # AD48
                "인원": {"row": 65, "col": 31},       # AE65
                "장비": {"row": 111, "col": 31}       # AE111
            }
        
        table_names = ["날씨정보", "시공현황", "작업내용", "인원", "장비"]
        
        for i, (table_name, df) in enumerate(zip(table_names, tables_data)):
            if df is not None and not df.empty:
                pos = table_positions.get(table_name, {"row": 10 + i*20, "col": 30})
                start_row = pos["row"]
                start_col = pos["col"]
                
                # 헤더 입력
                for col_idx, column_name in enumerate(df.columns):
                    ws.cell(row=start_row, column=start_col+col_idx, value=column_name)
                
                # 데이터 입력
                for row_idx, row_data in enumerate(df.itertuples(index=False)):
                    for col_idx, cell_value in enumerate(row_data):
                        ws.cell(row=start_row+1+row_idx, column=start_col+col_idx, value=cell_value)
        
        # 바이트 스트림으로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ 엑셀 파일 처리 중 오류 발생: {str(e)}")
        return None

# 페이지 설정
st.set_page_config(
    page_title="공사일보 자동화",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS 스타일 정의
st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
    }
    .title-text {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1E3A8A;
        margin-bottom: 0.5rem;
    }
    .subtitle-text {
        font-size: 1.2rem;
        color: #4B5563;
        margin-bottom: 2rem;
    }
    .step-container {
        display: flex;
        justify-content: space-between;
        margin: 2rem 0;
        position: relative;
    }
    .step-container::before {
        content: '';
        position: absolute;
        top: 20px;
        left: 0;
        right: 0;
        height: 2px;
        background: #E5E7EB;
        z-index: 1;
    }
    .step {
        display: flex;
        flex-direction: column;
        align-items: center;
        position: relative;
        z-index: 2;
    }
    .step-circle {
        width: 40px;
        height: 40px;
        border-radius: 50%;
        background: #E5E7EB;
        display: flex;
        align-items: center;
        justify-content: center;
        margin-bottom: 0.5rem;
        font-weight: 600;
        color: #4B5563;
    }
    .step.active .step-circle {
        background: #1E3A8A;
        color: white;
    }
    .step-label {
        font-size: 0.9rem;
        color: #4B5563;
        text-align: center;
    }
    .step.active .step-label {
        color: #1E3A8A;
        font-weight: 600;
    }
    .card {
        background: white;
        border-radius: 0.5rem;
        padding: 1.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        margin-bottom: 1.5rem;
    }
    .card-title {
        font-size: 1.25rem;
        font-weight: 600;
        color: #1E3A8A;
        margin-bottom: 1rem;
    }
    .card-description {
        color: #4B5563;
        margin-bottom: 1rem;
    }
    .stButton>button {
        background-color: #1E3A8A;
        color: white;
        border: none;
        padding: 0.5rem 1rem;
        border-radius: 0.375rem;
        font-weight: 500;
        transition: all 0.2s;
    }
    .stButton>button:hover {
        background-color: #1E40AF;
        transform: translateY(-1px);
    }
    .stTextArea>div>div>textarea {
        border-radius: 0.375rem;
        border: 1px solid #E5E7EB;
        padding: 0.75rem;
    }
    .stTextArea>div>div>textarea:focus {
        border-color: #1E3A8A;
        box-shadow: 0 0 0 2px rgba(30,58,138,0.1);
    }
    
    /* 프롬프트 수정 모달 스타일 */
    .modal {
        display: none;
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background-color: rgba(0,0,0,0.5);
        z-index: 1000;
    }
    .modal-content {
        position: relative;
        background-color: white;
        margin: 5% auto;
        padding: 2rem;
        width: 80%;
        max-width: 800px;
        border-radius: 0.5rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .modal-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 1rem;
    }
    .modal-title {
        font-size: 1.5rem;
        font-weight: 600;
        color: #1E3A8A;
    }
    .close-button {
        background: none;
        border: none;
        font-size: 1.5rem;
        cursor: pointer;
        color: #4B5563;
    }
    .button-container {
        display: flex;
        gap: 1rem;
        margin-top: 1rem;
    }
    .stTextArea>div>div>textarea {
        min-height: 400px;
        font-family: monospace;
    }
    .prompt-editor {
        background-color: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 0.375rem;
        padding: 1rem;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# 제목 및 부제목
st.markdown('<div class="title-text">공사일보 자동화</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle-text">AI 기반 작업보고서 생성 시스템</div>', unsafe_allow_html=True)

# 단계 표시
st.markdown("""
<div class="step-container">
    <div class="step active">
        <div class="step-circle">1</div>
        <div class="step-label">텍스트 입력</div>
    </div>
    <div class="step">
        <div class="step-circle">2</div>
        <div class="step-label">파일 업로드</div>
    </div>
    <div class="step">
        <div class="step-circle">3</div>
        <div class="step-label">미리보기</div>
    </div>
    <div class="step">
        <div class="step-circle">4</div>
        <div class="step-label">승인 및 다운로드</div>
    </div>
</div>
""", unsafe_allow_html=True)

# 최초 실행 시 세션 상태에 프롬프트 값이 없으면 기본값으로 초기화
if "prompt" not in st.session_state:
    st.session_state["prompt"] = DEFAULT_PROMPT

# 작업 완료 상태 추적을 위한 세션 상태 초기화
if "kakao_work_completed" not in st.session_state:
    st.session_state["kakao_work_completed"] = False
if "blast_data_completed" not in st.session_state:
    st.session_state["blast_data_completed"] = False
if "processed_tables" not in st.session_state:
    st.session_state["processed_tables"] = []
if "blast_dataframe" not in st.session_state:
    st.session_state["blast_dataframe"] = None

# 텍스트 입력 카드
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">카카오톡 작업보고 입력</div>', unsafe_allow_html=True)
st.markdown('<div class="card-description">일일작업보고 텍스트</div>', unsafe_allow_html=True)

# 텍스트 입력 영역
kakao_text = st.text_area(
    "",
    placeholder="카카오톡에서 공유된 작업보고 내용을 붙여넣으세요...",
    height=200
)

# 버튼 컨테이너
col1, col2 = st.columns([1, 1])

# AI로 구조화하기 버튼
with col1:
    if st.button("AI로 구조화하기", key="structure_button"):
        if kakao_text:
            try:
                prompt = st.session_state["prompt"] + "\n" + kakao_text
                response = GEMINI_MODEL.generate_content(prompt)
                preview_content = response.text
                tables = preview_content.split("```")
                table_names = ["날씨정보", "시공현황", "작업내용", "인원", "장비"]
                real_tables = []
                processed_tables = []  # DataFrame 형태로 저장
                
                for table in tables:
                    tsv_candidate = table.strip()
                    # 탭이 있으면 일단 테이블로 간주 (조건 완화)
                    if "\t" in tsv_candidate:
                        real_tables.append(tsv_candidate)
                
                for i, tsv_data in enumerate(real_tables):
                    tsv_data = remove_tsv_label(tsv_data)
                    tsv_data_fixed = fix_tsv_field_count(tsv_data)  # 필드 개수 보정
                    if i < len(table_names):
                        st.subheader(f"{table_names[i]} 테이블")
                    else:
                        st.subheader(f"테이블 {i+1}")
                    df = parse_tsv_to_dataframe(tsv_data_fixed)
                    if df is not None:
                        st.dataframe(df.reset_index(drop=True))
                        processed_tables.append(df)  # DataFrame 저장
                    else:
                        st.warning(f"테이블 {i+1}의 데이터를 파싱할 수 없습니다.")
                        processed_tables.append(None)
                
                st.download_button(
                    label="전체 미리보기 다운로드",
                    data=preview_content,
                    file_name="공사일보_미리보기.txt",
                    mime="text/plain"
                )
                
                # 5개 테이블이 모두 생성된 경우 완료 상태로 저장
                if len(processed_tables) >= 5 and any(df is not None for df in processed_tables):
                    st.session_state["kakao_work_completed"] = True
                    st.session_state["processed_tables"] = processed_tables
                    st.success("✅ 카카오톡 작업보고 처리 완료!")
                    st.info("📋 생성된 테이블: 날씨정보, 시공현황, 작업내용, 인원, 장비")
                    
                    # 입력된 테이블 정보
                    valid_tables = [name for i, name in enumerate(table_names) if i < len(processed_tables) and processed_tables[i] is not None]
                    st.info(f"📊 처리된 테이블: {', '.join(valid_tables)}")
                    
                    # 진행상황 안내
                    if not st.session_state["blast_data_completed"]:
                        st.info("⏳ 발파데이터 입력이 완료되면 샘플파일 업로드가 가능합니다.")
            except Exception as e:
                st.error(f"미리보기 생성 중 오류가 발생했습니다: {str(e)}")
                st.error("상세 오류 정보:")
                st.error(str(e))
        else:
            st.warning("카카오톡 작업보고 내용을 입력해주세요.")

# 프롬프트 수정 버튼
with col2:
    if st.button("프롬프트 수정", key="edit_prompt_button"):
        st.session_state.show_prompt_modal = True

# 프롬프트 수정 모달 (Streamlit 위젯만 사용)
if st.session_state.get('show_prompt_modal', False):
    st.markdown("### 프롬프트 수정", unsafe_allow_html=True)
    edited_prompt = st.text_area(
        "프롬프트를 수정하세요",
        value=st.session_state["prompt"],
        height=400,
        key="prompt_editor"
    )
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("저장", key="save_prompt_button"):
            st.session_state["prompt"] = edited_prompt
            st.session_state.show_prompt_modal = False
            st.experimental_rerun()
    with col2:
        if st.button("취소", key="cancel_prompt_button"):
            st.session_state.show_prompt_modal = False
            st.experimental_rerun()
    with col3:
        if st.button("닫기(X)", key="close_prompt_button"):
            st.session_state.show_prompt_modal = False
            st.experimental_rerun()

st.markdown('</div>', unsafe_allow_html=True)

# 파일 업로드 카드 (단일 업로더만 유지)
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown('<div class="card-title">파일 업로드</div>', unsafe_allow_html=True)

# 발파데이터 파일 업로드 (단일 업로더)
blast_files = st.file_uploader(
    "발파데이터 파일을 업로드하세요 (PDF, 엑셀, 워드 지원)",
    type=["pdf", "xlsx", "xls", "docx"],
    accept_multiple_files=True
)

# 파일 유형 판별 함수 추가
def identify_blast_files(uploaded_files):
    """업로드된 파일들을 분석하여 발파작업일지와 계측결과 보고서를 구분하는 함수"""
    blast_log_file = None
    daily_report_file = None
    
    # 파일명 기반 키워드 매칭
    blast_log_keywords = ["발파", "작업", "일지", "blast", "work", "log"]
    measurement_keywords = ["계측", "진동", "소음", "보고서", "measurement", "vibration", "noise", "report"]
    
    for uploaded_file in uploaded_files:
        filename_lower = uploaded_file.name.lower()
        
        # 발파작업일지 키워드 검사
        if any(keyword in filename_lower for keyword in blast_log_keywords):
            if blast_log_file is None:  # 첫 번째로 발견된 것만
                blast_log_file = uploaded_file
                continue
        
        # 계측결과 보고서 키워드 검사
        if any(keyword in filename_lower for keyword in measurement_keywords):
            if daily_report_file is None:  # 첫 번째로 발견된 것만
                daily_report_file = uploaded_file
                continue
    
    # 키워드로 구분이 안되면 업로드 순서대로 할당
    if blast_log_file is None or daily_report_file is None:
        if len(uploaded_files) >= 2:
            if blast_log_file is None:
                blast_log_file = uploaded_files[0]
            if daily_report_file is None:
                daily_report_file = uploaded_files[1] if uploaded_files[1] != blast_log_file else None
    
    return blast_log_file, daily_report_file

if blast_files and len(blast_files) == 2:
    # 파일 유형 자동 판별
    blast_log_file, daily_report_file = identify_blast_files(blast_files)
    
    if blast_log_file and daily_report_file:
        st.info(f"📄 발파작업일지: {blast_log_file.name}")
        st.info(f"📊 계측결과 보고서: {daily_report_file.name}")
        
        with st.spinner('🤖 AI가 데이터를 분석하고 있습니다...'):
            try:
                # 각 파일 내용 추출
                blast_text = extract_file_content(blast_log_file)
                daily_text = extract_file_content(daily_report_file)
                
                # 디버깅: 추출된 텍스트 확인
                with st.expander("📄 발파작업일지 원본 TSV (디버깅)"):
                    st.text_area("발파작업일지 TSV", blast_text if blast_text else "내용 없음", height=150)
                with st.expander("📊 계측일지 원본 TSV (디버깅)"):
                    st.text_area("계측일지 TSV", daily_text if daily_text else "내용 없음", height=150)
                
                if not blast_text or not daily_text:
                    st.error("❌ 파일 내용 추출에 실패하여 처리를 중단합니다.")
                else:
                    prompt = BLAST_EXTRACTION_PROMPT + f"\n\n## 입력 1: 발파작업일지_TSV\n{blast_text}\n\n## 입력 2: 계측일지_TSV\n{daily_text}"
                    
                    # 디버깅: 최종 프롬프트 확인
                    with st.expander("🔍 최종 프롬프트 (디버깅)"):
                        st.text_area("LLM 전달 프롬프트", prompt, height=300)
                    
                    response = GEMINI_MODEL.generate_content(prompt)
                    
                    # 디버깅: AI 응답 원본 확인
                    with st.expander("🤖 AI 응답 원본 (디버깅)"):
                        st.text_area("AI 응답", response.text if response.text else "응답 없음", height=200)
                    
                    tsv_result = response.text
                    tsv_data = extract_tsv_from_response(tsv_result)
                    
                    # 디버깅: extract_tsv_from_response 결과 확인
                    with st.expander("✂️ TSV 추출 결과 (디버깅)"):
                        st.text_area("추출된 TSV 문자열", tsv_data, height=200)
                    
                    if not tsv_data or not '\t' in tsv_data:
                        st.error("❌ AI 응답에서 유효한 TSV 데이터를 추출하지 못했습니다.")
                        st.info("팁: AI 응답 원본과 TSV 추출 결과(디버깅용)를 확인하여, BLAST_EXTRACTION_PROMPT 또는 extract_tsv_from_response 함수를 조정해보세요.")
                    else:
                        tsv_data = fix_tsv_field_count(tsv_data)  # 필드 개수 보정
                        
                        # 추가 검증 및 정제
                        tsv_data = validate_and_clean_tsv(tsv_data)
                        
                        # 디버깅: 보정된 TSV 데이터 확인
                        with st.expander("🔧 보정된 TSV 데이터 (디버깅)"):
                            st.text_area("보정된 TSV 문자열", tsv_data, height=200)
                        
                        try:
                            # TSV 파싱 시도 (최신 pandas 버전용)
                            df = pd.read_csv(
                                io.StringIO(tsv_data), 
                                sep='\t', 
                                encoding='utf-8',
                                on_bad_lines='skip',  # 문제가 있는 라인 건너뛰기
                                engine='python'  # python 엔진 사용으로 오류 처리 개선
                            )
                        except Exception as csv_error:
                            st.warning(f"⚠️ 최신 pandas 옵션 실패, 대체 방법 시도: {csv_error}")
                            try:
                                # 구버전 pandas 또는 대체 방법
                                df = pd.read_csv(
                                    io.StringIO(tsv_data), 
                                    sep='\t', 
                                    encoding='utf-8',
                                    error_bad_lines=False,  # 구버전용
                                    warn_bad_lines=True,
                                    engine='python'
                                )
                            except Exception as csv_error2:
                                st.error(f"❌ TSV 파싱 실패: {csv_error2}")
                                # 수동으로 TSV 파싱 시도
                                lines = tsv_data.strip().split('\n')
                                if lines:
                                    headers = lines[0].split('\t')
                                    data_rows = []
                                    for line in lines[1:]:
                                        row = line.split('\t')
                                        # 컬럼 수 맞추기
                                        if len(row) < len(headers):
                                            row += [''] * (len(headers) - len(row))
                                        elif len(row) > len(headers):
                                            row = row[:len(headers)]
                                        data_rows.append(row)
                                    df = pd.DataFrame(data_rows, columns=headers)
                                else:
                                    raise Exception("TSV 데이터가 비어있습니다.")
                        
                        df = df.fillna('-')
                        
                        st.success("✅ 발파데이터 분석 완료!")
                        st.subheader("📋 추출된 발파데이터")
                        st.dataframe(df.reset_index(drop=True))
                        st.download_button(
                            label="📥 TSV 파일로 다운로드",
                            data=df.to_csv(sep='\t', index=False, encoding='utf-8-sig'),
                            file_name="발파데이터.tsv",
                            mime="text/tab-separated-values",
                        )
                        
                        # 발파데이터 완료 상태로 저장
                        st.session_state["blast_data_completed"] = True
                        st.session_state["blast_dataframe"] = df
                        st.success("✅ 발파데이터 처리 완료!")
                        st.info(f"📊 처리된 데이터: {len(df)}행 × {len(df.columns)}열")
                        
                        # 진행상황 안내
                        if not st.session_state["kakao_work_completed"]:
                            st.info("⏳ 카카오톡 작업보고 입력이 완료되면 샘플파일 업로드가 가능합니다.")
            except Exception as e:
                st.error(f"❌ 데이터 분석 중 오류가 발생했습니다: {e}")
                st.error(f"상세 오류: {str(e)}")
    else:
        st.warning("⚠️ 파일 유형을 자동으로 구분할 수 없습니다.")
        st.info("파일명에 다음 키워드가 포함되어야 합니다:")
        st.info("• 발파작업일지: '발파', '작업', '일지' 중 하나")
        st.info("• 계측결과 보고서: '계측', '진동', '소음', '보고서' 중 하나")
elif blast_files and len(blast_files) == 1:
    st.info("📁 파일 1개가 업로드되었습니다. 발파데이터 분석을 위해서는 2개 파일이 필요합니다.")
elif blast_files and len(blast_files) > 2:
    st.warning("⚠️ 파일이 너무 많습니다. 발파작업일지와 계측결과 보고서 2개 파일만 업로드해주세요.")
else:
    st.info("📁 발파데이터 분석을 위해 관련 파일 2개를 업로드해주세요.")

# 사용 가능한 Gemini 모델 목록을 sidebar에 출력 (개발/디버깅용)
with st.sidebar:
    st.markdown('### 사용 가능한 Gemini 모델 목록')
    try:
        models = genai.list_models()
        for m in models:
            st.write(f"- {m.name}")
            if hasattr(m, 'supported_generation_methods'):
                st.caption(f"지원 메서드: {getattr(m, 'supported_generation_methods', None)}")
    except Exception as e:
        st.error(f"모델 목록 조회 실패: {str(e)}") 

# === AE160 자동 입력 기능 업데이트 필요 ===
# template_file 업로드 시 자동으로 AE160에 데이터 입력하도록 수정 필요

st.markdown('</div>', unsafe_allow_html=True)

# 통합 샘플파일 업로드 섹션 (두 작업이 모두 완료되었을 때만 표시)
if st.session_state["kakao_work_completed"] and st.session_state["blast_data_completed"]:
    st.markdown("---")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📊 샘플 엑셀 파일에 데이터 통합 입력</div>', unsafe_allow_html=True)
    st.markdown('<div class="card-description">처리된 모든 데이터를 샘플 엑셀 파일의 지정된 위치에 자동으로 입력합니다.</div>', unsafe_allow_html=True)
    
    # 완료된 작업 상태 표시
    st.info("✅ 카카오톡 작업보고 처리 완료")
    st.info("✅ 발파데이터 처리 완료")
    
    # 샘플 파일 업로드
    template_file = st.file_uploader(
        "샘플 엑셀 파일을 업로드하세요",
        type=["xlsx", "xls"],
        key="final_template_uploader"
    )
    
    if template_file:
        st.info("📍 데이터 입력 위치:")
        st.info("• 날씨정보: AD5 | 시공현황: AD13 | 작업내용: AD48 | 인원: AE65 | 장비: AE111")
        st.info("• 발파데이터: AE160")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📝 모든 데이터 입력하기", key="integrate_all_data"):
                with st.spinner('📝 엑셀 파일에 모든 데이터를 자동으로 입력하고 있습니다...'):
                    try:
                        # 1. 먼저 5개 테이블 입력
                        modified_excel = insert_five_tables_to_excel(
                            st.session_state["processed_tables"], 
                            template_file
                        )
                        
                        if modified_excel:
                            # 2. 임시 파일로 저장 후 발파데이터 추가
                            temp_file = io.BytesIO(modified_excel)
                            final_excel = insert_blast_data_to_excel_ae160(
                                st.session_state["blast_dataframe"], 
                                temp_file
                            )
                            
                            if final_excel:
                                st.success("✅ 모든 데이터 입력 완료!")
                                
                                # 수정된 파일 다운로드
                                original_name = template_file.name
                                name_without_ext = original_name.rsplit('.', 1)[0]
                                extension = original_name.rsplit('.', 1)[1] if '.' in original_name else 'xlsx'
                                new_filename = f"{name_without_ext}_통합데이터입력.{extension}"
                                
                                st.download_button(
                                    label="📥 통합 엑셀 파일 다운로드",
                                    data=final_excel,
                                    file_name=new_filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                
                                # 입력된 데이터 정보
                                st.info("📊 입력된 데이터:")
                                st.info(f"• 5개 테이블: 날씨정보, 시공현황, 작업내용, 인원, 장비")
                                st.info(f"• 발파데이터: {len(st.session_state['blast_dataframe'])}행 × {len(st.session_state['blast_dataframe'].columns)}열")
                            else:
                                st.error("❌ 발파데이터 입력 중 오류가 발생했습니다.")
                        else:
                            st.error("❌ 5개 테이블 입력 중 오류가 발생했습니다.")
                    except Exception as e:
                        st.error(f"❌ 데이터 통합 입력 중 오류 발생: {str(e)}")
        
        with col2:
            if st.button("🔄 작업 초기화", key="reset_all_work"):
                st.session_state["kakao_work_completed"] = False
                st.session_state["blast_data_completed"] = False
                st.session_state["processed_tables"] = []
                st.session_state["blast_dataframe"] = None
                st.success("✅ 모든 작업이 초기화되었습니다.")
                st.experimental_rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

elif st.session_state["kakao_work_completed"] or st.session_state["blast_data_completed"]:
    st.markdown("---")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📋 작업 진행 상황</div>', unsafe_allow_html=True)
    
    if st.session_state["kakao_work_completed"]:
        st.info("✅ 카카오톡 작업보고 처리 완료")
    else:
        st.warning("⏳ 카카오톡 작업보고 입력 대기 중")
    
    if st.session_state["blast_data_completed"]:
        st.info("✅ 발파데이터 처리 완료")
    else:
        st.warning("⏳ 발파데이터 입력 대기 중")
    
    st.info("💡 두 작업이 모두 완료되면 샘플파일 업로드가 가능합니다.")
    
    if st.button("🔄 작업 초기화", key="reset_partial_work"):
        st.session_state["kakao_work_completed"] = False
        st.session_state["blast_data_completed"] = False
        st.session_state["processed_tables"] = []
        st.session_state["blast_dataframe"] = None
        st.success("✅ 모든 작업이 초기화되었습니다.")
        st.experimental_rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

# 사용 가능한 Gemini 모델 목록을 sidebar에 출력 (개발/디버깅용)