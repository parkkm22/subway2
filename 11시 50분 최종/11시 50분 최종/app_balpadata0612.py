import streamlit as st
import pandas as pd
from datetime import datetime
import os
import google.generativeai as genai
import io
import re
import pdfplumber
from pdf2image import convert_from_bytes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import requests
import json

# --- CONFIG & SETUP ---
st.set_page_config(
    page_title="공사일보 자동화",
    page_icon="🏗️",
    layout="wide"
)

# --- STYLING ---
st.markdown("""
<style>
    /* Main App Font */
    html, body, [class*="st-"], .stTextArea, .stButton>button, .stFileUploader, .stSelectbox {
        font-family: 'Pretendard', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    }
    /* Main container */
    .main .block-container {
        padding: 2rem 2rem 5rem 2rem;
        max-width: 1000px;
    }
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #F8F9FA;
        border-right: 1px solid #E5E7EB;
    }
    [data-testid="stSidebar"] h1 {
        font-size: 1.5rem;
        color: #1E3A8A;
        font-weight: 700;
        padding: 1rem 0;
    }
    /* Step container in sidebar */
    .step-container {
        padding-top: 1rem;
    }
    .step {
        display: flex;
        align-items: center;
        margin-bottom: 1.25rem;
        padding: 0.75rem;
        border-radius: 0.5rem;
        transition: background-color 0.3s, border-color 0.3s;
        border-left: 5px solid #E5E7EB;
    }
    .step.active {
        border-left-color: #2563EB;
        background-color: #EFF6FF;
    }
    .step.completed {
        border-left-color: #16A34A;
    }
    .step-icon {
        font-size: 1.5rem;
        margin-right: 1rem;
    }
    .step-text {
        font-size: 1rem;
        font-weight: 500;
        color: #374151;
    }
    .step.completed .step-text {
        color: #115E59;
    }
    /* Main content cards */
    .card {
        background-color: white;
        border-radius: 0.75rem;
        padding: 2rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -2px rgba(0, 0, 0, 0.05);
        border: 1px solid #E5E7EB;
        margin-bottom: 2rem;
    }
    .card-title {
        font-size: 1.75rem;
        font-weight: 700;
        color: #1E3A8A;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
    }
    .card-title .icon {
        font-size: 2rem;
        margin-right: 0.75rem;
    }
    .card-description {
        color: #4B5563;
        margin-bottom: 1.5rem;
    }
    /* Custom button style */
    .stButton>button {
        background-color: #2563EB;
        color: white;
        border: none;
        padding: 0.75rem 1.5rem;
        border-radius: 0.5rem;
        font-weight: 600;
        width: 100%;
        transition: background-color 0.3s;
    }
    .stButton>button:hover {
        background-color: #1D4ED8;
    }
    .stButton>button:disabled {
        background-color: #9CA3AF;
        color: #E5E7EB;
    }
    .stButton>button.reset-button {
        background-color: #D1D5DB;
        color: #4B5563;
    }
    .stButton>button.reset-button:hover {
        background-color: #9CA3AF;
        color: #1F2937;
    }
</style>
""", unsafe_allow_html=True)


# --- GLOBAL CONSTANTS & API SETUP ---
TEAMS_WEBHOOK_URL = "https://poscoenc365.webhook.office.com/webhookb2/f6efcf11-c6a7-4385-903f-f3fd8937de55@ec1d3aa9-13ec-4dc5-8672-06fc64ca7701/IncomingWebhook/1fb9d9ce7f4c4093ba4fe9a8db67dc2f/1a2e3f7d-551b-40ec-90a1-e815373c81a7/V2qbqRtbAap4il8cvVljyk_ApZuHTDE0AfOYLQ8V9SqQs1"
GENAI_API_KEY = "AIzaSyD69-wKYfZSID327fczrkx-JveJdGYIUIk"
genai.configure(api_key=GENAI_API_KEY)
GEMINI_MODEL = genai.GenerativeModel("models/gemini-2.5-flash-preview-05-20")

BLAST_EXTRACTION_PROMPT = '''
# INSTRUCTION
- 반드시 아래 예시처럼 오직 TSV(탭 구분) 데이터만 출력하세요.
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 아래 예시와 동일한 형식으로만 출력하세요.
발파일자	발파시간	지발당장약량(최소, kg)	지발당장약량(최대, kg)	폭약사용량(kg)	발파진동(cm/sec)	발파소음(dB(A))	계측위치	비고
2023-07-27	08:05	0.5	0.9	73	-	-	-	PLA-2
2023-07-27	13:47	0.4	0.8	77	0.87	53.29	티스테이션	PD-2
2023-07-27	13:47	-	-	-	0.71	61.23	양말집	PD-2
(위 예시는 형식만 참고, 실제 데이터는 입력값에 따라 동적으로 생성)
# 입력
- 입력1: 발파작업일지_TSV (아래와 같은 형식)
- 입력2: 계측일지_TSV (아래와 같은 형식, **계측일지 표는 PDF 2페이지 이후부터 추출**)
# 입력1 예시
발파일자	발파시간	지발당장약량(최소, kg)	지발당장약량(최대, kg)	폭약사용량(kg)	비고
2023-07-27	08:05	0.5	0.9	73	PLA-2
2023-07-27	13:47	0.4	0.8	77	PD-2
# 입력2 예시 (**2페이지 이후 표만**)
Date/Time	Peak Particle Vel (X_Axis) (mm/sec)	Peak Particle Vel (Y_Axis) (mm/sec)	Peak Particle Vel (Z_Axis) (mm/sec)	LMax (Sound) (dBA)	측정위치
2023/07/27 1:47:00 PM	0.71	0.36	0.71	61.23	양말집
2023/07/27 1:47:00 PM	0.87	0.56	0.87	53.29	티스테이션
# Mapping Rules
- 두 입력을 병합하여 위 예시와 동일한 TSV만 출력
- 설명, 마크다운, 코드블록, 주석, 기타 텍스트는 절대 포함하지 마세요.
- 계측일지 표는 반드시 PDF 2페이지 이후의 표만 사용 
- 최종 헤더(고정열): 발파일자, 발파시간, 지발당장약량(최소, kg), 지발당장약량(최대, kg), 폭약사용량(kg), 발파진동(cm/sec), 발파소음(dB(A)), 계측위치, 비고
- 정렬: 발파시간 오름차순, 계측위치 오름차순(필요시)
- 병합/매칭/포맷 규칙은 기존과 동일
'''
DEFAULT_PROMPT = """
# INSTRUCTIONS
1. 기상청 서울 지역 관측 자료를 기반으로 "날씨정보" 테이블을 TSV(UTF-8) 형식의 별도 코드블록으로 생성
2. 일일작업보고 원문에서 데이터를 파싱하여 4개 테이블("시공현황", "작업내용", "인원", "장비") 각각을 TSV(UTF-8) 형식의 별도 코드블록으로 차례대로 출력하며 아래의 조건을 철저히 준수할 것
# OUTPUT : 테이블(총 5개)  
## 1. 날씨정보 테이블
1. 고정 열 : "구분", "값"
2. 고정 행 : "최고온도", "최저온도", "강수량"
3. 추출데이터 : 서울(유) 오늘 날씨 예보 (최신 업데이트)
4. 주의사항 
- 서울 지역(영등포구 우선)의 최고 기온, 최저 기온, 강수량의 단일값 추출
- 데이터는 최신 업데이트된 기상청 정보를 기반으로 제공
- "값"만 숫자로 추출할 것 (예: 20.0 °C에서 "20.0" 추출)
## 2. 시공현황 테이블  
1. 고정 열 : "구분", "누계"  
2. 고정 행(총 33행) - 아래 순서와 명칭을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)  
- "1. 본선터널 (1구간, 대림-신풍) 라이닝" 
- "2. 신풍정거장 - 1)정거장 라이닝"
- "2. 신풍정거장 - 1)정거장 미들 슬라브"
- "2. 신풍정거장 – 2)주출입구 수직구 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 정거장 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (1)PCB 환승통로 방면 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (2)PCC 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (3)PCD 라이닝"
- "2. 신풍정거장 - 2)주출입구 - (4)PHA 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - 수직구 라이닝"
- "2. 신풍정거장 - 3)특별피난계단 - PHB 라이닝"
- "2. 신풍정거장 - 4)외부출입구 출입구(#3) 굴착" 
- "2. 신풍정거장 - 4)외부출입구 출입구(#2) 굴착"
- "2. 신풍정거장 - 4)외부출입구 출입구(#1) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCF) 라이닝"  
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 굴착" 
- "3. 신풍 환승통로 - 1)환승터널 연결터널(PCE) 라이닝"  
- "3. 신풍 환승통로 - 2)개착 BOX 보라매 방면 구조물"  
- "3. 신풍 환승통로 - 2)개착 BOX 대림 방면 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 굴착"  
- "4. 본선터널(2구간, 신풍-도림) 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 터널 라이닝"  
- "5. 도림사거리정거장 - 1)정거장 미들 슬라브" 
- "5. 도림사거리정거장 - 2)출입구#1 수직구 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCA 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PCC 라이닝"  
- "5. 도림사거리정거장 - 2)출입구#1 PHA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 수직구 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCA 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PCC 라이닝"  
- "5. 도림사거리정거장 - 3)출입구#2 PHB 라이닝"  
3. 추출데이터  
- "누계"값만 숫자로 추출할 것 (예: 945.3m / 1,116m 에서 "945.3" 추출)
## 3. 작업내용 테이블  
1. 고정 열 : "구분", "금일작업"  
2. 고정 행(총 14행) - 아래 순서와 명칭(매핑 후 결과)을 그대로  
- "1. 본선터널 (1구간, 대림-신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍-도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"  
3. 주의사항  
- '작업내용' 셀은 여러 세부 내용을 포함할 수 있습니다. 내용을 구분할 때는, 최종 TSV 출력 시 해당 셀을 큰따옴표("...")로 감싸되, 셀 내부의 각 내용은 **실제 줄바꿈 문자(예: '\\n' 문자열 대신 엔터 키 입력에 해당)**를 사용하여 분리하며, '-'기호는 생략함
## 4. 인원 / 장비 테이블  
1. 고정 열 (총 15열) - 열 순서는 아래와 같음
- "구분" 
- "1. 본선터널 (1구간, 대림~신풍)"  
- "2.신풍정거장 - 1)정거장 터널"  
- "2.신풍정거장 - 2)주출입구 - (1)PCB"  
- "2.신풍정거장 - 2)주출입구 - (2)PCC"  
- "2.신풍정거장 - 2)주출입구 - (3)PCD"  
- "2.신풍정거장 - 2)주출입구 - (4)PHA"  
- "2.신풍정거장 - 3)특별피난계단"  
- "2.신풍정거장 - 4)외부출입구"  
- "3.신풍 환승통로 - 1)환승터널"  
- "3.신풍 환승통로 - 2)개착 BOX"  
- "4.본선터널(2구간, 신풍~도림)"  
- "5.도림사거리정거장 - 1)정거장 터널"  
- "5.도림사거리정거장 - 2)출입구#1"  
- "5.도림사거리정거장 - 3)출입구#2"    
2. 고정 행(인원 테이블 – 총 36행)  
(인원 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
"직영반장", "연수생", "장비운전원", "전기주임", "화약주임", "터널공", "목공", "철근공", "라이닝폼공", "오폐수처리공", "카리프트공", "BP공", "가시설공/해체공", "동바리공", "신호수", "부단수공", "슬러리월공", "CIP공", "미장공", "시설물공", "경계석공", "조경공", "배관공", "도색공", "방수공", "장비/작업지킴이", "보통인부", "포장공", "용접공", "타설공", "보링공/앙카공", "비계공", "도장공", "석면공", "주입공/그라우팅공"
3. 고정 행 (장비 테이블 – 총 46행)  
(장비 목록은 아래 순서와 명칭(매핑 후 결과)을 반드시 그대로 사용):
"B/H(1.0LC)", "B/H(08W)", "B/H(08LC)", "B/H(06W)", "B/H(06LC)", "B/H(03LC)", "B/H(02LC)", "B/H(015)", "덤프트럭(5T)", "덤프트럭(15T)", "덤프트럭(25T)", "앵글크레인(100T)", "앵글크레인(80T)", "앵글크레인(35T)", "앵글크레인(25T)", "카고크레인(25T)", "카고크레인(5T)", "콤프", "점보드릴", "페이로더", "숏트머신", "차징카", "살수차", "하이드로크레인", "믹서트럭", "화물차(5T)", "펌프카", "스카이", "콘크리트피니셔", "전주오거", "로더(바브켓)", "유제살포기(비우다)", "지게차", "싸인카", "BC커터기", "바이브로해머", "롤러(2.5T)", "롤러(1T)", "롤러(0.7T)", "몰리", "항타기", "크레인", "콤비로라", "공압드릴", "유압드릴", "기타"
## 5. Parsing Rules 
1. 시공현황: "누계/설계" → **앞 값(소수 허용)** 만 추출.    
2. 인원·장비: 투입현황에서 **정수만** 추출, 빈셀은 **0**    
3. 하위 섹션 매핑    
   - 정거장 터널 → 열 ②, PCB → ③, PCC → ④, PCD → ⑤, PHA → ⑥, 특별피난 → ⑦, 외부출입구 → ⑧    
4. 매핑 딕셔너리 적용    
- "B/H08W" → "B/H(08W)"   
- "25톤 카고크레인" → "카고크레인(25T)"   
- "특공" → "보통인부"    
- "기계타설공" → "타설공"    
- "목공연수생" 또는 "목수연수생" → "연수생"    
- "5톤트럭" → "화물차(5T)"    
- "카리프트" → "카리프트공"    
- "하이드로크레인(20T)" → "하이드로크레인"    
- "라이닝폼조립" → "라이닝폼공"  
- "S/C타설팀" → "터널공"  
- "목수" → "목공"    
5. 사전에 없는 항목 → 유사항목, 없으면 **인원: 보통인부 / 장비: 기타** 로 합산하고 '오류요약'에 기재.
## 6. 변환로그 (변경사항이 있을 때만 출력)
변경된 항목만 아래 형식으로 출력:
(원문) 목수 -> (변경) 목공   *위치: 1. 본선터널(1구간, 대림-신풍)
(원문) 특공 -> (변경) 보통인부   *위치: 2.신풍정거장 - 1)정거장 터널
(원문) B/H08W -> (변경) B/H(08W)   *위치: 4.본선터널(2구간, 신풍-도림)
주의사항:
- 변경사항이 없으면 "변환로그: 변경사항 없음" 출력
- 각 변경사항은 별도 행으로 출력
- 위치는 구체적인 작업 구간명 기재
"""

# --- HELPER FUNCTIONS ---
def safe_generate_content(model_input):
    """
    Calls the Gemini API with robust error handling and relaxed safety settings.
    """
    try:
        # AI 모델의 안전 설정을 완화하여 콘텐츠 차단을 최소화합니다.
        safety_settings = {
            'HARM_CATEGORY_HARASSMENT': 'BLOCK_NONE',
            'HARM_CATEGORY_HATE_SPEECH': 'BLOCK_NONE',
            'HARM_CATEGORY_SEXUALLY_EXPLICIT': 'BLOCK_NONE',
            'HARM_CATEGORY_DANGEROUS_CONTENT': 'BLOCK_NONE',
        }
        
        response = GEMINI_MODEL.generate_content(
            model_input,
            safety_settings=safety_settings
        )

        # 응답에 실제 콘텐츠(parts)가 있는지 확인합니다.
        if response.parts:
            return response.text
        else:
            # 콘텐츠가 없는 경우, 차단 원인을 확인하여 사용자에게 알립니다.
            reason = "Unknown"
            try:
                # API 응답에서 제공하는 공식적인 차단 이유를 가져옵니다.
                reason = response.prompt_feedback.block_reason.name
            except Exception:
                pass 
            st.error(f"AI 응답 생성에 실패했습니다. API에 의해 콘텐츠가 차단되었을 수 있습니다. (차단 이유: {reason})")
            st.warning(f"전체 피드백: {response.prompt_feedback}")
            return None
            
    except Exception as e:
        st.error(f"AI 모델 호출 중 심각한 오류 발생: {e}")
        return None

def send_teams_alert(warning_rows, file_date):
    try:
        message = {
            "type": "message",
            "attachments": [{
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": {
                    "type": "AdaptiveCard",
                    "body": [
                        {"type": "TextBlock", "size": "Large", "weight": "Bolder", "text": f"⚠️ 계측기 경고 알림 ({file_date})", "color": "Attention"},
                        {"type": "TextBlock", "text": "다음 계측기에서 주의가 필요한 변화가 감지되었습니다:", "wrap": True}
                    ]
                }
            }]
        }
        for _, row in warning_rows.iterrows():
            warning_info = {"type": "TextBlock", "text": f"📍 위치: {row['위치']}\\n\\n📊 계측기: {row['계측기명']} ({row['계측기 종류']})\\n\\n⚠️ 상태: {row['상태']}\\n\\n📈 3차 초과 대비: {row['비율']}", "wrap": True, "style": "warning"}
            message["attachments"][0]["content"]["body"].append(warning_info)
        
        response = requests.post(TEAMS_WEBHOOK_URL, json=message, headers={"Content-Type": "application/json"})
        if response.status_code == 200: st.success("Teams로 경고 메시지가 전송되었습니다.")
        else: st.error(f"Teams 메시지 전송 실패: {response.status_code}")
    except Exception as e: st.error(f"Teams 메시지 전송 중 오류 발생: {e}")

def extract_file_content(file):
    if file.name.endswith('.pdf'):
        try:
            file.seek(0)
            uploaded_file = genai.upload_file(file, mime_type="application/pdf")
            
            filename_lower = file.name.lower()
            is_measurement_file = any(keyword in filename_lower for keyword in ["계측", "진동", "소음"])
            is_blast_log_file = any(keyword in filename_lower for keyword in ["발파", "작업", "일지"])

            if is_measurement_file:
                pdf_prompt = "이 PDF 파일은 '발파진동소음 계측일지'입니다. 다음 지침에 따라 데이터를 TSV 형식으로 추출해주세요. ... (Prompt content is long and omitted for brevity)"
            elif is_blast_log_file:
                pdf_prompt = "이 PDF 파일은 '발파작업일지'입니다. 다음 지침에 따라 주요 데이터를 TSV 형식으로 추출해주세요. ... (Prompt content is long and omitted for brevity)"
            else:
                st.warning("⚠️ 파일 유형을 특정할 수 없어 일반 표 추출을 시도합니다.")
                pdf_prompt = "이 PDF에서 가장 중요해 보이는 표를 찾아 TSV 형식으로 추출해주세요. ..."

            # 안전하게 AI 모델을 호출합니다.
            response_text = safe_generate_content([pdf_prompt, uploaded_file])
            
            # 사용이 끝난 파일은 즉시 삭제합니다.
            genai.delete_file(uploaded_file.name)

            if response_text:
                return re.sub(r'```tsv|```', '', response_text).strip()
            
            return None # safe_generate_content에서 오류를 이미 표시했으므로 None만 반환합니다.

        except Exception as e:
            st.error(f"❌ {file.name} 처리 중 AI 오류 발생: {e}")
            return None
    elif file.name.endswith(('.xlsx', '.xls')):
        try:
            return pd.read_excel(file, engine='openpyxl').to_csv(sep='\t', index=False, encoding='utf-8')
        except Exception as e:
            st.error(f"❌ 엑셀 데이터 추출 실패: {e}")
            return None
    return None

def parse_tsv_to_dataframe(tsv_content):
    try:
        cleaned_content = '\n'.join(line.strip() for line in tsv_content.split('\n') if line.strip())
        df = pd.read_csv(io.StringIO(cleaned_content), sep='\t', encoding='utf-8')
        df.columns = df.columns.str.strip()
        return df.fillna('')
    except Exception as e:
        st.error(f"TSV 파싱 중 오류 발생: {e}")
        return None

def extract_tsv_from_response(response_text):
    if not response_text: return ""
    lines = response_text.strip().split('\n')
    cleaned_lines = [line.strip() for line in lines if '\t' in line.strip()]
    return "\n".join(cleaned_lines)

def fix_tsv_field_count(tsv_str):
    lines = tsv_str.strip().split('\n')
    if not lines: return tsv_str
    
    header = lines[0]
    n_fields = header.count('\t') + 1
    fixed_lines = [header]
    
    for line in lines[1:]:
        fields = line.split('\t')
        if len(fields) < n_fields:
            fields += [''] * (n_fields - len(fields))
        elif len(fields) > n_fields:
            fields = fields[:n_fields-1] + [' '.join(fields[n_fields-1:])]
        fixed_lines.append('\t'.join(fields))
    
    return '\n'.join(fixed_lines)

def insert_data_to_excel(template_bytes, **kwargs):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    
    # 5 Tables
    tables_data = kwargs.get("tables_data")
    if tables_data:
        table_positions = {
            "날씨정보": {"row": 4, "col": 30}, "시공현황": {"row": 12, "col": 30},
            "작업내용": {"row": 47, "col": 30}, "인원": {"row": 64, "col": 31},
            "장비": {"row": 110, "col": 31}
        }
        for name, df in zip(["날씨정보", "시공현황", "작업내용", "인원", "장비"], tables_data):
            if df is not None:
                pos = table_positions[name]
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    for c_idx, value in enumerate(row):
                        ws.cell(row=pos["row"] + r_idx, column=pos["col"] + c_idx, value=value)

    # Blast Data
    blast_df = kwargs.get("blast_df")
    if blast_df is not None:
        for r_idx, row in enumerate(dataframe_to_rows(blast_df, index=False, header=False)):
            for c_idx, value in enumerate(row):
                ws.cell(row=160 + r_idx, column=31 + c_idx, value=value)

    # Instrument Data
    instrument_df = kwargs.get("instrument_df")
    if instrument_df is not None:
        for r_idx, row in enumerate(dataframe_to_rows(instrument_df, index=False, header=False)):
            for c_idx, value in enumerate(row):
                ws.cell(row=171 + r_idx, column=31 + c_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# --- STATE INITIALIZATION ---
def initialize_session_state():
    states = {
        "kakao_work_completed": False, "blast_data_completed": False,
        "instrument_data_completed": False, "excel_export_completed": False,
        "processed_tables": [], "blast_dataframe": None,
        "instrument_dataframe": None, "kakao_results": None,
        "final_excel_data": None, "processed_template_filename": None,
        "all_accumulated_rows": [], "reset_flag": 0,
        "prompt": DEFAULT_PROMPT,
        "warning_rows_instrument": None
    }
    for key, value in states.items():
        if key not in st.session_state:
            st.session_state[key] = value

initialize_session_state()


# --- SIDEBAR ---
with st.sidebar:
    st.markdown("<h1>🏗️ 공사일보 자동화</h1>", unsafe_allow_html=True)
    
    if st.button("모든 작업 초기화", key="reset_all", use_container_width=True, type="secondary"):
        for key in list(st.session_state.keys()): del st.session_state[key]
        st.rerun()

    st.markdown('<div class="step-container">', unsafe_allow_html=True)

    steps = [
        ("카카오톡 작업보고 입력", "kakao_work_completed", "📝"),
        ("발파 데이터 확인", "blast_data_completed", "🧨"),
        ("자동화 계측기 데이터 확인", "instrument_data_completed", "📈"),
        ("공사일보 엑셀 추출", "excel_export_completed", "📄")
    ]
    
    current_step_index = 0
    if st.session_state.kakao_work_completed: current_step_index = 1
    if st.session_state.blast_data_completed: current_step_index = 2
    if st.session_state.instrument_data_completed: current_step_index = 3
    if st.session_state.excel_export_completed: current_step_index = 4

    for i, (text, state_key, icon) in enumerate(steps):
        is_completed = st.session_state.get(state_key, False)
        is_active = (i == current_step_index)
        status_class = "completed" if is_completed else "active" if is_active else ""
        step_icon = "✅" if is_completed else icon
        
        st.markdown(f'<div class="step {status_class}"><div class="step-icon">{step_icon}</div><div class="step-text"><strong>{i+1}.</strong> {text}</div></div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# --- MAIN CONTENT ---
st.title("공사일보 자동화 시스템")
st.markdown("AI 기반 작업보고서 생성 및 데이터 통합 솔루션. 사이드바의 단계를 따라 진행해주세요.")
st.markdown("---")

# --- STEP 1: KAKAO TALK REPORT ---
with st.container():
    st.markdown('<div class="card"><div class="card-title"><span class="icon">📝</span>1. 카카오톡 작업보고 입력</div><p class="card-description">카카오톡 내용을 붙여넣고 AI로 구조화하여 데이터를 추출합니다.</p>', unsafe_allow_html=True)
    
    if not st.session_state.kakao_work_completed:
        kakao_text = st.text_area("일일작업보고 텍스트", placeholder="이곳에 카카오톡 작업보고 내용을 붙여넣으세요...", height=200, label_visibility="collapsed")
        if st.button("AI로 구조화하기", key="structure_button", use_container_width=True):
            if kakao_text:
                with st.spinner('🤖 AI가 데이터를 분석하고 있습니다...'):
                    try:
                        prompt = st.session_state.prompt + "\n" + kakao_text
                        response_text = safe_generate_content(prompt)
                        
                        if response_text:
                            st.session_state.kakao_results = response_text
                            st.session_state.kakao_work_completed = True
                            st.toast("✅ 1단계 완료: 카카오톡 데이터 구조화 성공!")
                            st.rerun()
                        # 'else'의 경우 safe_generate_content 함수 내부에서 오류 메시지가 이미 표시됨
                    except Exception as e: 
                        st.error(f"미리보기 생성 중 오류: {e}")
            else: 
                st.warning("보고 내용을 입력해주세요.")
    else:
        st.success("✅ 1단계 완료: 카카오톡 작업보고가 성공적으로 처리되었습니다.")
        with st.expander("처리된 테이블 데이터 보기"):
            tables = st.session_state.kakao_results.split("```")
            table_names = ["날씨정보", "시공현황", "작업내용", "인원", "장비"]
            real_tables = [t.strip() for t in tables if "\t" in t.strip()]
            
            processed_tables = []
            for i, tsv_data in enumerate(real_tables):
                df = parse_tsv_to_dataframe(fix_tsv_field_count(re.sub(r'^tsv\n', '', tsv_data, flags=re.IGNORECASE)))
                if df is not None:
                    st.subheader(table_names[i] if i < len(table_names) else f"테이블 {i+1}")
                    st.dataframe(df)
                    processed_tables.append(df)
            st.session_state.processed_tables = processed_tables
    st.markdown('</div>', unsafe_allow_html=True)

# --- STEP 2: BLASTING DATA ---
if st.session_state.kakao_work_completed:
    with st.container():
        st.markdown('<div class="card"><div class="card-title"><span class="icon">🧨</span>2. 발파 데이터 확인</div><p class="card-description">발파작업일지와 계측결과 보고서를 업로드하여 데이터를 병합하고 정제합니다.</p>', unsafe_allow_html=True)
        
        if not st.session_state.blast_data_completed:
            blast_files = st.file_uploader("발파작업일지 및 계측결과 보고서 (2개 파일)", type=["pdf", "xlsx", "xls"], accept_multiple_files=True, key=f"blast_files_{st.session_state.reset_flag}")
            
            if len(blast_files) == 2:
                with st.spinner('🤖 AI가 발파 데이터를 분석하고 있습니다...'):
                    try:
                        blast_text = extract_file_content(blast_files[0])
                        daily_text = extract_file_content(blast_files[1])
                        
                        if blast_text and daily_text:
                            prompt = BLAST_EXTRACTION_PROMPT + f"\n\n## 입력 1: 발파작업일지_TSV\n{blast_text}\n\n## 입력 2: 계측일지_TSV\n{daily_text}"
                            response_text = safe_generate_content(prompt)

                            if response_text:
                                tsv_result = extract_tsv_from_response(response_text)
                                df = parse_tsv_to_dataframe(fix_tsv_field_count(tsv_result))
                                
                                if df is not None:
                                    st.session_state.blast_dataframe = df
                                    st.session_state.blast_data_completed = True
                                    st.toast("✅ 2단계 완료: 발파 데이터 분석 성공!")
                                    st.rerun()
                                else: 
                                    st.error("AI 응답에서 유효한 TSV를 추출하지 못했습니다.")
                            # 'else'의 경우 safe_generate_content 함수 내부에서 오류 메시지가 이미 표시됨
                        else: 
                            st.error("파일 내용 추출에 실패했습니다.")
                    except Exception as e: 
                        st.error(f"데이터 분석 중 오류: {e}")
        else:
            st.success("✅ 2단계 완료: 발파 데이터가 성공적으로 처리되었습니다.")
            with st.expander("처리된 발파 데이터 보기"):
                st.dataframe(st.session_state.blast_dataframe)
        st.markdown('</div>', unsafe_allow_html=True)

# --- STEP 3: INSTRUMENT DATA ---
if st.session_state.blast_data_completed:
    with st.container():
        st.markdown('<div class="card"><div class="card-title"><span class="icon">📈</span>3. 자동화 계측기 데이터 확인</div><p class="card-description">계측기 엑셀 파일을 업로드하여 최대 변화량을 분석합니다.</p>', unsafe_allow_html=True)

        if not st.session_state.instrument_data_completed:
            excel_files = st.file_uploader("자동화 계측기 엑셀 파일(들)", type=["xlsx", "xls"], accept_multiple_files=True, key=f"inst_files_{st.session_state.reset_flag}")
            
            if excel_files:
                with st.spinner("🔄 자동화 계측기 데이터를 처리하는 중입니다..."):
                    try:
                        all_accumulated_rows = []
                        for uploaded_file in excel_files:
                            try:
                                xls = pd.ExcelFile(uploaded_file)
                                for sheet_name in xls.sheet_names:
                                    df = pd.read_excel(xls, sheet_name=sheet_name)
                                    if df.empty or df.shape[0] < 2:
                                        continue

                                    first_row_values = df.columns.tolist()
                                    data_values = df.values.tolist()
                                    last_row_values = data_values[-1] if data_values else []
                                    
                                    location_val = sheet_name.replace("ALL", "").replace("all", "").replace("All", "").strip()
                                    location_val = " ".join(location_val.split()) # 중복 공백 제거

                                    if "주출입구" in location_val:
                                        location_val = "신풍 주출입구"
                                    elif "단면" in location_val:
                                        location_val = "신풍 특피"
                                    elif "출입구" in location_val and not location_val.startswith("도림"):
                                        location_val = location_val.replace("출입구", "도림출입구")

                                    if isinstance(last_row_values, (list, tuple)) and len(last_row_values) > 1:
                                        for col_idx, current_value in enumerate(last_row_values[1:], 1):
                                            if col_idx < len(first_row_values):
                                                instrument_name = str(first_row_values[col_idx])
                                                current_value_str = str(current_value)
                                                weekly_change = "-"
                                                try:
                                                    if len(data_values) >= 2:
                                                        last_val = float(str(data_values[-1][col_idx]))
                                                        first_val = float(str(data_values[0][col_idx]))
                                                        weekly_change = str(round(last_val - first_val, 3))
                                                except (ValueError, IndexError):
                                                    weekly_change = "-"
                                                
                                                if current_value_str.lower() != "nan":
                                                    all_accumulated_rows.append([
                                                        location_val, instrument_name, weekly_change, current_value_str
                                                    ])
                            except Exception as e:
                                st.warning(f"'{uploaded_file.name}' 처리 중 오류: {e}")

                        if all_accumulated_rows:
                            temp_df = pd.DataFrame(all_accumulated_rows, columns=["위치", "계측기명", "주간변화량", "누적변화량"])
                            temp_df['계측기 종류'] = temp_df['계측기명'].apply(lambda x: 
                                "변형률계" if "변형률" in str(x) else
                                "지하수위계" if "W" in str(x) or "지하수위" in str(x) else
                                "지중경사계" if "INC" in str(x) or "지중경사" in str(x) else
                                "ST하중계" if "하중" in str(x) else "기타")
                            temp_df['단위'] = temp_df['계측기명'].apply(lambda x: 
                                "ton" if "변형률" in str(x) or "하중" in str(x) else
                                "m" if "W" in str(x) or "지하수위" in str(x) else
                                "mm" if "INC" in str(x) or "지중경사" in str(x) else "")
                            
                            temp_df = temp_df[temp_df['계측기 종류'] != '기타']
                            
                            # "하중계"는 '계측기명'이 'R'로 끝나는 것만 필터링
                            is_st_load_cell = temp_df['계측기 종류'] == "ST하중계"
                            ends_with_r = temp_df['계측기명'].str.strip().str.upper().str.endswith('R')
                            # ST하중계가 아니거나, ST하중계이면서 R로 끝나는 경우만 유지
                            temp_df = temp_df[~is_st_load_cell | (is_st_load_cell & ends_with_r)]

                            summary_df = temp_df[temp_df["주간변화량"] != "-"].copy()
                            if not summary_df.empty:
                                summary_df["주간변화량_float"] = pd.to_numeric(summary_df["주간변화량"], errors='coerce').fillna(0)
                                summary_df["누적변화량_float"] = pd.to_numeric(summary_df["누적변화량"], errors='coerce').fillna(0)
                                summary_df["주간변화량_절대값"] = summary_df["주간변화량_float"].abs()
                                
                                # 1. 최대 변화량 데이터부터 요약
                                max_changes = summary_df.loc[summary_df.groupby(["위치", "계측기 종류"])["주간변화량_절대값"].idxmax()].copy()

                                def determine_status(row):
                                    try:
                                        if row['계측기 종류'] == "ST하중계":
                                            value = abs(row['누적변화량_float'])
                                            limit = 100
                                            if value >= limit: return "3차 초과", value/limit
                                            elif value >= limit*0.8: return "2차 초과", value/limit
                                            elif value >= limit*0.6: return "1차 초과", value/limit
                                            else: return "안정", value/limit
                                        elif row['계측기 종류'] == "변형률계":
                                            value = abs(row['누적변화량_float'])
                                            limit = 2518
                                            if value >= limit: return "3차 초과", value/limit
                                            elif value >= limit*0.8: return "2차 초과", value/limit
                                            elif value >= limit*0.6: return "1차 초과", value/limit
                                            else: return "안정", value/limit
                                        elif row['계측기 종류'] == "지중경사계":
                                            value = abs(row['누적변화량_float'])
                                            limit = 128.96
                                            if value >= limit: return "3차 초과", value/limit
                                            elif value >= limit*0.8: return "2차 초과", value/limit
                                            elif value >= limit*0.6: return "1차 초과", value/limit
                                            else: return "안정", value/limit
                                        elif row['계측기 종류'] == "지하수위계":
                                            value = abs(row['주간변화량_float'])
                                            limit = 1.0
                                            if value >= limit: return "3차 초과", value/limit
                                            elif value >= limit*0.75: return "2차 초과", value/limit
                                            elif value >= limit*0.5: return "1차 초과", value/limit
                                            else: return "안정", value/limit
                                        return "확인필요", 0
                                    except (ValueError, TypeError): return "오류", 0

                                # 2. 요약된 데이터에 대해서만 상태 분석 수행
                                status_results = max_changes.apply(determine_status, axis=1)
                                max_changes['상태'] = status_results.apply(lambda x: x[0])
                                max_changes['비율'] = status_results.apply(lambda x: f"{x[1]*100:.1f}%" if x[1] > 0 else "N/A")
                                max_changes['누적변화량'] = max_changes['누적변화량_float'].apply(lambda x: f"{x:.3f}")
                                
                                # 3. 화면 표시용과 엑셀 저장용 데이터프레임 모두 요약본 기반으로 생성
                                display_df = max_changes[["위치", "계측기 종류", "계측기명", "주간변화량", "누적변화량", "단위", "상태", "비율"]]
                                excel_export_df = max_changes[['위치', '계측기 종류', '계측기명', '주간변화량', '누적변화량', '단위', '상태']].copy()

                                # 4. 두 데이터프레임을 각각 세션에 저장 (엑셀용은 요약본)
                                st.session_state['instrument_display_df'] = display_df
                                st.session_state['instrument_dataframe'] = excel_export_df
                                st.session_state.instrument_data_completed = True
                                
                                # 경고 알림은 요약본(최대값) 기준으로 찾아 세션에 저장만 함
                                warning_rows = display_df[display_df['상태'].str.contains("초과")]
                                st.session_state['warning_rows_instrument'] = warning_rows

                                st.toast("✅ 3단계 완료: 자동화 계측기 데이터 분석 성공!")
                                st.rerun()

                    except Exception as e:
                        st.error(f"데이터 분석 중 오류 발생: {e}")

        else:
            st.success("✅ 3단계 완료: 자동화 계측기 데이터가 성공적으로 처리되었습니다.")
            with st.expander("최대 변화량 분석 결과 보기"):
                if 'instrument_display_df' in st.session_state and not st.session_state.instrument_display_df.empty:
                    df_to_display = st.session_state.instrument_display_df

                    def highlight_warning_rows(row):
                        if row['상태'] != '안정':
                            return ['background-color: #ffcdd2'] * len(row)
                        return [''] * len(row)
                    
                    styled_df = df_to_display.style.apply(highlight_warning_rows, axis=1)
                    st.dataframe(styled_df)

                    # Teams 알림 수동 전송 버튼
                    warning_rows = st.session_state.get('warning_rows_instrument')
                    if warning_rows is not None and not warning_rows.empty:
                        st.warning(f"🚨 {len(warning_rows)}개의 항목에서 관리기준 초과가 감지되었습니다.")
                        if st.button("⚠️ Teams로 경고 메시지 전송", key="send_teams_alert_manual"):
                            file_date_for_alert = datetime.now().strftime('%y%m%d')
                            send_teams_alert(warning_rows, file_date_for_alert)
                else:
                    st.info("표시할 분석 결과가 없습니다.")
        st.markdown('</div>', unsafe_allow_html=True)


# --- STEP 4: EXCEL EXPORT ---
if st.session_state.instrument_data_completed:
    with st.container():
        st.markdown('<div class="card"><div class="card-title"><span class="icon">📄</span>4. 공사일보 엑셀 추출</div><p class="card-description">모든 데이터를 통합하여 최종 공사일보를 생성합니다.</p>', unsafe_allow_html=True)

        if not st.session_state.excel_export_completed:
            template_file = st.file_uploader("샘플 엑셀 템플릿 파일 업로드", type=["xlsx", "xls"], key=f"template_file_{st.session_state.reset_flag}")
            if template_file:
                with st.spinner("🚀 최종 엑셀 파일을 생성하고 있습니다..."):
                    try:
                        template_bytes = template_file.read()
                        final_excel_bytes = insert_data_to_excel(
                            template_bytes,
                            tables_data=st.session_state.processed_tables,
                            blast_df=st.session_state.blast_dataframe,
                            instrument_df=st.session_state.instrument_dataframe
                        )
                        
                        original_name = template_file.name
                        new_filename = f"{original_name.rsplit('.', 1)[0]}_통합완료.xlsx"
                        
                        st.session_state.final_excel_data = {'data': final_excel_bytes, 'filename': new_filename}
                        st.session_state.excel_export_completed = True
                        st.toast("🎉 모든 작업 완료! 최종 파일을 다운로드하세요.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"최종 파일 생성 중 오류: {e}")

        if st.session_state.excel_export_completed:
            st.success("🎉 최종 공사일보가 성공적으로 생성되었습니다!")
            final_data = st.session_state.final_excel_data
            st.download_button(
                label="📥 최종 공사일보 다운로드",
                data=final_data['data'],
                file_name=final_data['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        st.markdown('</div>', unsafe_allow_html=True)
